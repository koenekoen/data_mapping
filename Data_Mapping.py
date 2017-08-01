"""
|-------------------------------------|
|  Data Mapping for Quintiq MP Tool   |
|-------------------------------------|

New in v2.4:
    Quintiq
        Updated taxonomy conversions
        Reworked some outputs to debugging/missing
        Added GCMF POs
        Fixed some feedback/capacity bugs related to open POs
    Optimus
        N/A
    General
        N/A

To be added:
    Quintiq
        Long-term Agreements
            [m] FSA - tricky because they hold globally, not for a specific RB - waiting on data from Planners
            [l] Shipping - check with shipping if data exists
        Load transactional data
            [m] Shipments - included in LESS High Seas stock? yes, but doesn't show lead time!
        Storage Costs
            [m] Specify by location (currently by commodity)
        Bulk shipments
            [l] Lead times - LESS data @ RBD is very young, not much data to build on
        GCMF
            [m] Refine costs - currently last confirmed (PO) price, by specific commodity (not location)
            [l] Add GCMF project explicitly (incl. purchase options) ??? cost of handover between GCMF and project would become nonlinear?
            [m] Alternative design: preference bonus for GCMF hubs (ports, warehouses)
            [l] Split up GCMF by sub-zone
        BBD
            [l] Currently using 6 months default
        Lanes
            [l] Some empty lanes (product-wise) for local procurement connections
            [h] Establish CO-CO links!
        In-Kind
            [l] Macros may break if multiple entries for the same donor & commodity combination
    Optimus
        Everything
    General
        Create a GUI?

File versions:
    QID: v2.3
    QMP: v5.3.2.1
    Model: 0.0.720.0
    OPT: v0.9.0.1

"""

# Import external packages
from __future__ import division # always true division when using /
import time # Allows the tracking of time
import os # Allows access to (sub)folders
import datetime # Allows conversion of excel's ridiculous date format
import calendar
import openpyxl as xl
import math
import difflib
from shutil import copyfile



class MP_Mapping:
    def __init__(self):
        # Initialize
        self.script_dir = os.path.dirname(os.path.abspath(__file__))
        self.data_dir = os.path.join(self.script_dir,"Bare Necessities")
        self.dest_dir = os.path.join(self.script_dir,"Python MP")

        # Load data
        t_0 = time.time()
        #self.db = xl.load_workbook(filename = "Master Data Repository.xlsx", read_only=True, data_only=True)
        # Comment Tim: For me, the data is only recognized when read_only=False. Maybe different version...
        self.db = xl.load_workbook(filename = "Master Data Repository.xlsx", read_only=True, data_only=True)
        print "<< Loading all data >>"
        print " "
        self.load_data_master()
        self.missing = [] # keep track of data mismatches
        self.load_data_conversions()
        self.load_data_resources()
        self.load_data_demand()
        self.load_data_sourcing()
        self.load_data_logistics()
        self.load_data_transactional()
        self.load_data_constraints()
        self.load_data_other()
        self.print_missing()

        self.db._archive.close()
        t_1 = time.time()
        print "<< Finished loading all data in ", "{0:.3f}".format(t_1-t_0), " seconds >>"
        print " "

        # Create GUI
        # self.draw_GUI(window)

        # Remove redundant connections
        print "<< Removing redundant data (Quintiq MP) >>"
        print " "
        self.Q_RB = "RBD"
        self.Q_Start_Horizon = datetime.datetime(2017,4,1)
        self.Q_Start_Planning = add_months(self.Q_Start_Horizon,2)
        self.GCMF_Priority = "Hard" # p in [None,Soft,Hard]
        #self.Q_COs = ['SOUTH SUDAN','ETHIOPIA','SOMALIA','KENYA','YEMEN'] # if [], self.Q_RB defines selection of salessegments
        self.Q_COs = []
        self.remove_bs()
        t_2 = time.time()
        print " "
        print "<< Finished cleaning data in ", "{0:.3f}".format(t_2-t_1), " seconds >>"
        print " "

        # debug pre-export
        print "<< Exporting data to Excel >>"
        self.print_to_file(self.GCMF_Commodity,"GCMF Commodity Mapping",["Country","Project","Commodity","Index"])
        demands = {}
        for (c,p,d,sk),(f,nf,t) in self.TacticalDemand.items():
            try:
                pipeline = self.PipelineDemand[c,p,d,sk]
            except:
                pipeline = (0,0,0)
            demands[c,p,d,sk] = (f,nf,t) + pipeline
        for (c,p,d,sk),(f,nf,t) in self.PipelineDemand.items():
            if (c,p,d,sk) in demands.keys():
                continue
            demands[c,p,d,sk] = (0,0,0,f,nf,t)
        header = ["Country","Project","Date","Commodity",
                  "Funded (Tactical)","Non-Funded (Tactical)","Implementation Plan (Tactical)",
                  "Implementation Plan (Pipeline)","Non-Funded (Pipeline)","Project Plan (Pipeline)"]
        self.print_to_file(demands,"Pipeline vs Tactical",header)
        self.print_to_file(self.Q_SpecificCommodities_S_NDP,"Availability by NDP",["NDP","Coms"])
        t_3 = time.time()
        print "<< Finished exporting in ", "{0:.3f}".format(t_3-t_2), " seconds >>"
        print " "

        # Export data
        print "<< Exporting data to Quintiq MP >>"
        print " "
        self.export_Quintiq()
        t_4 = time.time()
        print " "
        print "<< Finished exporting data to Quintiq MP in ", "{0:.3f}".format(t_4-t_3), " seconds >>"
        print " "

        # debug post-export
##        demand = []
##        for (c,p) in self.Q_SalesSegments:
##            for k in self.Q_Commodities_D_CO[c]:
##                sk = self.Q_Commodities_SpecCom[k]
##                try:
##                    s = self.Q_Supply[c,p,sk]
##                except:
##                    s = 0
##                td = 0
##                for t in range(12):
##                    try:
##                        d = add_months(self.Q_Start_Planning,t)
##                        td += self.TacticalDemand_CF[c,p,d,sk] + self.TacticalDemand_IK[c,p,d,sk]
##                    except:
##                        None
##                if (c,p,sk) not in self.Q_Supply.keys() and td == 0:
##                    continue
##                fcr = self.StaticFCR[c,p,sk]
##                budget = td * fcr
##                discount = s * fcr
##                bonus = (s-td)*fcr if s>td else 0
##                demand.append((c,p,sk,td,s,fcr,budget,discount,bonus))
##        header = ("Country","Project","Commodity","Total Demand","Total Supply","FCR Rate","Expected Budget","Supply Discount","Budget Bonus")
##        self.print_to_file(demand,"Demand vs Pipeline",header)

        # Wrapping up
        print "<< Finished mapping algorithm in ", "{0:.3f}".format(t_4-t_0), " seconds >>"
        print " "

    def load_data_master(self):
        '''
        Unfiltered loading of the master data
        '''

        def regional_bureaux():
            print "> Loading Regional Bureaux"
            t_s = time.time()
            ws = self.db['0.01 Regional Bureaux']
            M = []
            for row in ws:
                line = []
                for cell in row:
                    line.append(cell.value)
                M.append(line)
            self.MD_Regional_Bureaux = []
            for r in range(8,len(M)):
                v = M[r][1]
                if v == None:
                    continue
                else:
                    self.MD_Regional_Bureaux.append(v)
            self.MD_Regional_Bureaux.sort()
            t_e = time.time()
            print "Duration: ", "{0:.3f}".format(t_e-t_s), " seconds"

        def countries():
            print "> Loading Countries"
            t_s = time.time()
            ws = self.db['0.02 Countries']
            M = []
            for row in ws:
                line = []
                for cell in row:
                    line.append(cell.value)
                M.append(line)
            self.MD_Countries = []
            self.MD_Countries_RB = {}
            for r in range(8,len(M)):
                v = M[r][2]
                if v == None:
                    continue
                else:
                    self.MD_Countries.append(v)
                    self.MD_Countries_RB[v] = M[r][1]
            self.MD_Countries.sort()
            t_e = time.time()
            print "Duration: ", "{0:.3f}".format(t_e-t_s), " seconds"

        def projects():
            print "> Loading Projects"
            t_s = time.time()
            ws = self.db['0.03 Projects']
            M = []
            for row in ws:
                line = []
                for cell in row:
                    line.append(cell.value)
                M.append(line)
            self.MD_Projects = []
            self.MD_Projects_Type = {}
            self.MD_Projects_Countries = {}
            self.MD_Projects_RB = {}
            for r in range(8,len(M)):
                p = str(M[r][4])
                if p == None:
                    continue
                else:
                    self.MD_Projects.append(p)
                    self.MD_Projects_RB[p] = M[r][1]
                    if p not in self.MD_Projects_Countries.keys():
                        self.MD_Projects_Countries[p]= []
                    self.MD_Projects_Countries[p].append(M[r][2])
                    self.MD_Projects_Type[p] = M[r][3]
                    self.MD_Countries_RB[M[r][2]] = M[r][1]
            self.MD_Projects.sort()
            t_e = time.time()
            print "Duration: ", "{0:.3f}".format(t_e-t_s), " seconds"

        def locations():
            print "> Loading Locations"
            t_s = time.time()
            ws = self.db['0.04 Locations']
            M = []
            for row in ws:
                line = []
                for cell in row:
                    line.append(cell.value)
                M.append(line)
            self.MD_Locations = []
            self.MD_Locations_Type = {}
            self.MD_Locations_TypeID = {}
            self.MD_Locations_Country = {}
            self.MD_Locations_NDPs = []
            self.MD_Locations_COs = []
            self.MD_Locations_DPs = []
            self.MD_Locations_EDPs = []
            for r in range(8,len(M)):
                v = M[r][2]
                if v == None:
                    continue
                else:
                    self.MD_Locations.append(v)
                    self.MD_Locations_Type[v] = M[r][5]
                    self.MD_Locations_TypeID[v] = M[r][4]
                    self.MD_Locations_Country[v] = M[r][1]
                    l = self.MD_Locations_Type[v]
                    if l == "Named Delivery Place":
                        self.MD_Locations_NDPs.append(v)
                    elif l == "Country Office":
                        self.MD_Locations_COs.append(v)
                    elif l == "Discharge Port":
                        self.MD_Locations_DPs.append(v)
                    elif l == "Extended Delivery Point":
                        self.MD_Locations_EDPs.append(v)
                    else:
                        print "Location Type not recognised:", l
            self.MD_Locations.sort()
            self.MD_Locations_NDPs.sort()
            self.MD_Locations_COs.sort()
            self.MD_Locations_DPs.sort()
            self.MD_Locations_EDPs.sort()
            t_e = time.time()
            print "Duration: ", "{0:.3f}".format(t_e-t_s), " seconds"

        def commodities():
            print "> Loading Commodities"
            t_s = time.time()
            ws = self.db['1.07 Commodity Food Group']
            M = []
            for row in ws:
                line = []
                for cell in row:
                    line.append(cell.value)
                M.append(line)
            self.MD_Commodities = []
            self.MD_CommodityTypes = []
            self.MD_FoodGroups = []
            self.MD_Commodities_Type = {}
            self.MD_Commodities_Group = {}
            self.MD_CommodityTypes_Group = {}
            for r in range(8,len(M)):
                v = M[r][2]
                if v == None:
                    continue
                else:
                    self.MD_Commodities.append(v)
                    self.MD_CommodityTypes.append(M[r][1])
                    self.MD_FoodGroups.append(M[r][3])
                    self.MD_Commodities_Type[v] = M[r][1]
                    self.MD_Commodities_Group[v] = M[r][3]
                    self.MD_CommodityTypes_Group[M[r][1]] = M[r][3]
            self.MD_Commodities.sort()
            self.MD_CommodityTypes = list(set(self.MD_CommodityTypes))
            self.MD_CommodityTypes.sort()
            self.MD_FoodGroups = list(set(self.MD_FoodGroups))
            self.MD_FoodGroups.sort()
            t_e = time.time()
            print "Duration: ", "{0:.3f}".format(t_e-t_s), " seconds"

        def currencies():
            print "> Loading Currencies"
            t_s = time.time()
            ws = self.db['0.06 Currencies']
            M = []
            for row in ws:
                line = []
                for cell in row:
                    line.append(cell.value)
                M.append(line)
            self.MD_Currencies = []
            self.MD_Currencies_Symbol = {}
            for r in range(8,len(M)):
                v = M[r][1]
                if v == None:
                    continue
                else:
                    self.MD_Currencies.append(v)
                    self.MD_Currencies_Symbol[v] = M[r][2]
            self.MD_Currencies.sort()
            t_e = time.time()
            print "Duration: ", "{0:.3f}".format(t_e-t_s), " seconds"

        def shipping_types():
            print "> Loading Shipping Types"
            t_s = time.time()
            ws = self.db['0.17 Shipping Type']
            M = []
            for row in ws:
                line = []
                for cell in row:
                    line.append(cell.value)
                M.append(line)
            self.MD_ShippingTypes = []
            for r in range(8,len(M)):
                v = M[r][1]
                if v == None:
                    continue
                else:
                    self.MD_ShippingTypes.append(v)
            self.MD_ShippingTypes.sort()
            t_e = time.time()
            print "Duration: ", "{0:.3f}".format(t_e-t_s), " seconds"

        # Execute
        t_0 = time.time()
        print "Loading Master Data"
        regional_bureaux()
        countries()
        projects()
        locations()
        commodities()
        currencies()
        shipping_types()
        t_e = time.time()
        print "Finished loading master data in ", "{0:.3f}".format(t_e-t_0), " seconds"
        print " "

    def load_data_conversions(self):
        '''
        Unfiltered loading of the conversion tables
        '''

        def general():
            print "> Regional Bureaux"
            t_s = time.time()
            ws = self.db['RB Mapping']
            M = []
            for row in ws:
                line = []
                for cell in row:
                    line.append(cell.value)
                M.append(line)
            self.CONV_RB = {}
            for r in range(2,len(M)):
                rb = M[r][2]
                if rb == None:
                    continue
                elif rb not in self.MD_Regional_Bureaux:
                    self.missing.append(["RB",rb,"RB Mapping"])
                else:
                    self.CONV_RB[M[r][1]] = rb
            t_e = time.time()
            print "Duration: ", "{0:.3f}".format(t_e-t_s), " seconds"

            print "> Countries"
            t_s = time.time()
            ws = self.db['Country Mapping']
            M = []
            for row in ws:
                line = []
                for cell in row:
                    line.append(cell.value)
                M.append(line)
            self.CONV_Country = {}
            for r in range(2,len(M)):
                c = M[r][2]
                if c == None:
                    continue
                elif c not in self.MD_Countries:
                    self.missing.append(["Country",c,"Country Mapping"])
                    continue
                self.CONV_Country[M[r][1]] = c
            t_e = time.time()
            print "Duration: ", "{0:.3f}".format(t_e-t_s), " seconds"

            print "> Named Delivery Places"
            t_s = time.time()
            ws = self.db['NDP Mapping']
            M = []
            for row in ws:
                line = []
                for cell in row:
                    line.append(cell.value)
                M.append(line)
            self.CONV_NDP = {}
            for r in range(2,len(M)):
                ndp = M[r][2]
                if ndp == None:
                    continue
                elif ndp not in self.MD_Locations_NDPs and ndp != 0:
                    self.missing.append(["NDP",ndp,"NDP Mapping"])
                else:
                    self.CONV_NDP[M[r][1]] = ndp
            t_e = time.time()
            print "Duration: ", "{0:.3f}".format(t_e-t_s), " seconds"

            print "> Discharge Ports"
            t_s = time.time()
            ws = self.db['DP Mapping']
            M = []
            for row in ws:
                line = []
                for cell in row:
                    line.append(cell.value)
                M.append(line)
            self.CONV_DP = {}
            for r in range(2,len(M)):
                dp = M[r][2]
                if dp == None:
                    continue
                elif dp not in self.MD_Locations_DPs:
                    self.missing.append(["DP",dp,"DP Mapping"])
                else:
                    self.CONV_DP[M[r][1]] = dp
            t_e = time.time()
            print "Duration: ", "{0:.3f}".format(t_e-t_s), " seconds"

            print "> Commodities"
            t_s = time.time()
            ws = self.db['Commodity Mapping']
            M = []
            for row in ws:
                line = []
                for cell in row:
                    line.append(cell.value)
                M.append(line)
            self.CONV_Commodity = {}
            for r in range(2,len(M)):
                k = M[r][3]
                if k == None:
                    continue
                elif k not in self.MD_Commodities:
                    self.missing.append(["Commodity",k,"Commodity Mapping"])
                else:
                    self.CONV_Commodity[str(M[r][1]),M[r][2].upper()] = k
            t_e = time.time()
            print "Duration: ", "{0:.3f}".format(t_e-t_s), " seconds"


        def less():
            print "> LESS Conversion"
            t_s = time.time()
            ws = self.db['LESS - Loc']
            M = []
            for row in ws:
                line = []
                for cell in row:
                    line.append(cell.value)
                M.append(line)
            self.CONV_LESS_Location = {}
            for r in range(2,len(M)):
                v = M[r][3]
                if v == None:
                    continue
                else:
                    self.CONV_LESS_Location[M[r][1],M[r][2]] = v # [Rec. Country, Cur. Country] -> [Location]

            ws = self.db['LESS Location Link']
            M = []
            for row in ws:
                line = []
                for cell in row:
                    line.append(cell.value)
                M.append(line)
            self.CONV_LESS_LocationLink = {}
            self.GCMF_Ports = []
            for r in range(2,len(M)):
                sc = M[r][1]
                if sc == None:
                    continue
                qloc = M[r][2]
                self.CONV_LESS_LocationLink[sc] = qloc
                if M[r][3] == 1:
                    self.GCMF_Ports.append(qloc)
            t_e = time.time()
            print "Duration: ", "{0:.3f}".format(t_e-t_s), " seconds"

        def po():
            print "> PO Conversion"
            t_s = time.time()
            ws = self.db['PO - IPO']
            M = []
            for row in ws:
                line = []
                for cell in row:
                    line.append(cell.value)
                M.append(line)
            self.CONV_ZFPF = {}
            for r in range(2,len(M)):
                po = M[r][2]
                if po == None:
                    continue
                ipo = M[r][1]
                self.CONV_ZFPF[ipo] = po

            ws = self.db['PO - Origins']
            M = []
            for row in ws:
                line = []
                for cell in row:
                    line.append(cell.value)
                M.append(line)
            self.PO_Origins = {}
            for r in range(2,len(M)):
                po = M[r][1]
                if po == None:
                    continue
                oc = M[r][2]
                if oc not in self.MD_Countries:
                    self.missing.append(["Country",oc,"PO Origins"])
                    continue
                ndp = M[r][3]
                if ndp not in self.MD_Locations_NDPs:
                    self.missing.append(["NDP",ndp,"PO Origins"])
                    continue
                self.PO_Origins[po] = (oc,ndp)

            t_e = time.time()
            print "Duration: ", "{0:.3f}".format(t_e-t_s), " seconds"

        def gcmf():
            print "> GCMF Coverage"
            t_s = time.time()
            ws = self.db['GCMF Commodities']
            M = []
            for row in ws:
                line = []
                for cell in row:
                    line.append(cell.value)
                M.append(line)
            self.GCMF_Coverage = {}
            for r in range(2,len(M)):
                zone = M[r][1]
                if zone == None:
                    continue
                com = M[r][2]
                v = 1 if M[r][3]=="X" else 0
                self.GCMF_Coverage[zone,com] = v

            ws = self.db['GCMF Zones']
            M = []
            for row in ws:
                line = []
                for cell in row:
                    line.append(cell.value)
                M.append(line)
            self.GCMF_Zone = {}
            for r in range(2,len(M)):
                co = M[r][1].upper()
                if co == None:
                    continue
                elif co not in self.MD_Countries:
                    if co in self.CONV_Country.keys():
                        co = self.CONV_Country[co]
                    else:
                        self.missing.append(["Country Mapping",co,"GCMF Zones"])
                        continue
                self.GCMF_Zone[co] = M[r][2]

            print "> GCMF WBS"
            t_s = time.time()
            ws = self.db['GCMF WBS']
            M = []
            for row in ws:
                line = []
                for cell in row:
                    line.append(cell.value)
                M.append(line)
            self.GCMF_WBS = {}
            for r in range(2,len(M)):
                wbs = M[r][1]
                if wbs == None:
                    continue
                zone = M[r][2]
                rb = M[r][3]
                self.GCMF_WBS[wbs] = (zone, rb)

            t_e = time.time()
            print "Duration: ", "{0:.3f}".format(t_e-t_s), " seconds"

        # Execute
        t_0 = time.time()
        print "Loading Conversion Tables"
        general()
        less()
        po()
        gcmf()
        t_e = time.time()
        print "Finished loading conversion tables in ", "{0:.3f}".format(t_e-t_0), " seconds"
        print " "

    def load_data_demand(self):
        '''
        Unfiltered loading of the demand data
        '''

        def global_pipeline():
            print "> Loading Global Pipeline"
            t_s = time.time()
            ws = self.db['Global Pipeline']
            M = []
            for row in ws:
                line = []
                for cell in row:
                    line.append(cell.value)
                M.append(line)
            self.PipelineDemand= {}
            self.StaticFCR = {}
            self.GCMF_Commodity = {}
            for r in range(8,len(M)):
                # Verify master data integrity
                if M[r][6] <= 0:
                    continue
                c = M[r][2].upper() # country
                if c not in self.MD_Countries:
                    if c in self.CONV_Country.keys():
                        c = self.CONV_Country[c]
                    else:
                        self.missing.append(["Country",c,"Global Pipeline"])
                        continue
                if c in self.GCMF_Zone.keys():
                    zone = self.GCMF_Zone[c]
                else:
                    self.missing.append(["Zone Mapping",c,"Global Pipeline"])
                    continue
                p = str(M[r][3]) # project
                if p not in self.MD_Projects:
                    self.missing.append(["Project",p,"Global Pipeline"])
                    continue
                d0 = str(M[r][4])
                d = datetime.datetime(int(d0[:4]),int(d0[-2:]),1)
                k0 = M[r][5].upper() # commodity (Pipeline)
                if (p,k0) in self.CONV_Commodity.keys():
                    k = self.CONV_Commodity[p,k0] # commodity (SCIPS)
                elif ('*',k0) in self.CONV_Commodity.keys():
                    k = self.CONV_Commodity['*',k0] # default conversion for this commodity
                else:
                    self.missing.append(["Commodity Conversion",k0 + " @ " + c + "_" + p,"Global Pipeline"])
                    continue
                # Load info
                self.PipelineDemand[c,p,d,k] = (M[r][7] if M[r][7]>0 else 0,M[r][9] if M[r][9]>0 else 0,M[r][6] if M[r][6]>0 else 0)
                # NB: (Country, Project, Date, Commodity): (Implementation Plan, Non-Funded, Project Plan)
                self.StaticFCR[c,p,k] = M[r][8]
                if (c,p,k) not in self.GCMF_Commodity.keys():
                    try:
                        self.GCMF_Commodity[c,p,k] = self.GCMF_Coverage[zone,k0]
                    except:
                        self.missing.append(["Commodity Coverage",k0 + " @ " + zone,"Global Pipeline"])
                        self.GCMF_Commodity[c,p,k] = 0
            t_e = time.time()
            print "Duration: ", "{0:.3f}".format(t_e-t_s), " seconds"

        def tactical_demand():
            print "> Loading Tactical Demand Plan"
            t_s = time.time()
            ws = self.db['Tactical Demand']
            M = []
            for row in ws:
                line = []
                for cell in row:
                    line.append(cell.value)
                M.append(line)
            self.TacticalDemand = {}
            for r in range(8,len(M)):
                # Verify master data integrity
                req = M[r][7]
                if req == None or req <= 0:
                    continue
                zone = M[r][2]
                c = M[r][3].upper() # country
                if c not in self.MD_Countries:
                    if c in self.CONV_Country.keys():
                        c = self.CONV_Country[c]
                    else:
                        self.missing.append(["Country",c,"Tactical Demand"])
                        continue
                p = str(M[r][4]) # project
                if p not in self.MD_Projects:
                    self.missing.append(["Project",p,"Tactical Demand"])
                    continue
                k0 = M[r][5].upper() # commodity (Tactical Demand)
                if (p,k0) in self.CONV_Commodity.keys():
                    k = self.CONV_Commodity[p,k0] # commodity (SCIPS)
                elif ('*',k0) in self.CONV_Commodity.keys():
                    k = self.CONV_Commodity['*',k0] # default conversion for this commodity
                else:
                    self.missing.append(["Commodity Conversion",k0 + " @ " + c + "_" + p,"Tactical Demand"])
                    continue
                d = add_months(M[r][6],1) # date
                # Load info
                self.TacticalDemand[c,p,d,k] = (M[r][10] if M[r][10]>0 else 0,M[r][11] if M[r][11]>0 else 0,M[r][7] if M[r][7]>0 else 0)
                # NB: (Country, Project, Date, Commodity): (Funded, Non-Funded, Total)
                if (c,p,k) not in self.GCMF_Commodity.keys():
                    try:
                        self.GCMF_Commodity[c,p,k] = self.GCMF_Coverage[zone,k0]
                    except:
                        self.missing.append(["Commodity Coverage",k0 + " @ " + zone,"Tactical Demand"])
                        continue
                if (c,p,k) not in self.StaticFCR.keys():
                    if (c,p,k) in self.FCR.keys():
                        self.StaticFCR[c,p,k] = self.FCR[c,p,k]
                    elif (c,p) in self.FCR.keys():
                        self.StaticFCR[c,p,k] = self.FCR[c,p]
                        self.missing.append(["FCR Rate (com)",c + " @ " + p + " @ " + k,"Tactical Demand"])
                    else:
                        self.StaticFCR[c,p,k] = 1000 # WFP-wide approximation
                        self.missing.append(["FCR Rate (any)",c + " @ " + p + " @ " + k,"Tactical Demand"])
            t_e = time.time()
            print "Duration: ", "{0:.3f}".format(t_e-t_s), " seconds"

        # Execute
        t_0 = time.time()
        print "Loading Demand Data"
        global_pipeline()
        tactical_demand()
        t_e = time.time()
        print "Finished loading demand data in ", "{0:.3f}".format(t_e-t_0), " seconds"
        print " "

    def load_data_sourcing(self):
        '''
        Unfiltered loading of the sourcing data
        '''

        def price_food():
            print "> Loading Commodity Prices (Food)"
            t_s = time.time()
            ws = self.db['2.01 Commodity Prices (Food)']
            M = []
            for row in ws:
                line = []
                for cell in row:
                    line.append(cell.value)
                M.append(line)
            self.Procurement_Inco = {}
            self.Procurement_Cost = {}
            self.Procurement_Date = {}
            for r in range(8,len(M)):
                # Verify master data integrity
                oc = M[r][1] # Origin Country
                if oc == None:
                    continue
                elif oc not in self.MD_Countries:
                    self.missing.append(["Country",oc,"Commodity Prices (Food)"])
                    continue
                ndp = M[r][2] # Named Delivery Place
                if ndp not in self.MD_Locations_NDPs:
                    self.missing.append(["NDP",ndp,"Commodity Prices (Food)"])
                    continue
                com = M[r][5] # commodity
                if com not in self.MD_Commodities:
                    self.missing.append(["Commodity",com,"Commodity Prices (Food)"])
                    continue
                d = M[r][9]
                gmo = M[r][6]
                if gmo == "GMO":
                    gmo = 1
                else:
                    gmo = 0
                # Load info
                key = (oc,ndp,com,gmo,d)
                if key in self.Procurement_Cost.keys():
                    if self.Procurement_Cost[key] < M[r][8]: # use the cheapest incoterm if dates are equal
                        continue
                self.Procurement_Inco[key] = M[r][3]
                self.Procurement_Cost[key] = M[r][8]
                self.Procurement_Date[key] = M[r][10] # As Of Date! Last Updated Date is part of the key
            t_e = time.time()
            print "Duration: ", "{0:.3f}".format(t_e-t_s), " seconds"

        def price_forecast():
            print "> Loading Price Forecasts"
            t_s = time.time()
            ws = self.db['2.04 Price Forecast']
            M = []
            for row in ws:
                line = []
                for cell in row:
                    line.append(cell.value)
                M.append(line)
            self.Forecast = {}
            for r in range(8,len(M)):
                # Verify master data integrity
                oc = M[r][1] # Origin Country
                if oc == None:
                    continue
                elif oc not in self.MD_Countries:
                    self.missing.append(["Country",oc,"Price Forecasts"])
                    continue
                ndp = M[r][2] # Named Delivery Place
                if ndp not in self.MD_Locations_NDPs:
                    self.missing.append(["NDP",ndp,"Price Forecasts"])
                    continue
                com = M[r][3] # commodity
                if com not in self.MD_Commodities:
                    self.missing.append(["Commodity",com,"Price Forecasts"])
                    continue
                gmo = M[r][4] # gmo
                if gmo == "GMO":
                    gmo = 1
                else:
                    gmo = 0
                d = M[r][5] # date
                # Load info
                self.Forecast[oc,ndp,com,gmo,d] = M[r][6]
            t_e = time.time()
            print "Duration: ", "{0:.3f}".format(t_e-t_s), " seconds"

        def gmo():
            print "> Loading GMO Preferences"
            t_s = time.time()
            ws = self.db['2.05 GMO']
            M = []
            for row in ws:
                line = []
                for cell in row:
                    line.append(cell.value)
                M.append(line)
            self.GMO = {}
            for r in range(8,len(M)):
                # Verify master data integrity
                c = M[r][1] # Origin Country
                if c == None:
                    continue
                elif c not in self.MD_Countries:
                    self.missing.append(["Country",c,"GMO"])
                    continue
                gmo = M[r][2]
                # Load info
                self.GMO[c] = gmo
            t_e = time.time()
            print "Duration: ", "{0:.3f}".format(t_e-t_s), " seconds"

        def usik():
            print "> Commodity Prices (Donation)"
            t_s = time.time()
            ws = self.db['2.09 CommodityPrices (Donation)']
            M = []
            for row in ws:
                line = []
                for cell in row:
                    line.append(cell.value)
                M.append(line)
            self.DonatedCommodity_Cost = {}
            for r in range(8,len(M)):
                v = M[r][3]
                if v == None or v == "N/A":
                    continue
                donor = M[r][1]
                k0 = M[r][2]
                if k0 not in self.MD_Commodities:
                    if ("*",k0) in self.CONV_Commodity.keys():
                        sk = self.CONV_Commodity["*",k0]
                    else:
                        self.missing.append(["Commodity Conversion",k0,"Donation prices"])
                        continue
                else:
                    sk = k0
                # date = M[r][4]
                self.DonatedCommodity_Cost[donor,sk] = v
            t_e = time.time()
            print "Duration: ", "{0:.3f}".format(t_e-t_s), " seconds"

        # Execute
        t_0 = time.time()
        print "Loading Sourcing Data"
        price_food()
        price_forecast()
        gmo()
        #usik() -> loaded indirectly from PO now
        t_e = time.time()
        print "Finished loading sourcing data in ", "{0:.3f}".format(t_e-t_0), " seconds"
        print " "

    def load_data_logistics(self):
        '''
        Unfiltered loading of the logistics data
        '''

        def container_rates():
            print "> Loading Container Rates"
            t_s = time.time()
            ws = self.db['3.01 Container Rates']
            M = []
            for row in ws:
                line = []
                for cell in row:
                    line.append(cell.value)
                M.append(line)
            self.Shipping_Rate = {}
            self.Shipping_Date = {}
            for r in range(8,len(M)):
                # Verify master data integrity
                lp = M[r][2] # Load Port
                if lp == None:
                    continue
                elif lp not in self.MD_Locations_NDPs:
                    self.missing.append(["Named Delivery Place",lp,"Container Rates"])
                    continue
                dp = M[r][4] # Discharge Port
                if dp not in self.MD_Locations_DPs:
                    self.missing.append(["Discharge Port",dp,"Container Rates"])
                    continue
                # Load info
                key = (lp,dp)
                self.Shipping_Rate[key] = M[r][7]
                self.Shipping_Date[key] = M[r][6]
            t_e = time.time()
            print "Duration: ", "{0:.3f}".format(t_e-t_s), " seconds"

        def commodity_intakes():
            print "> Loading Commodity Intakes"
            t_s = time.time()
            ws = self.db['3.03b Commodity Intake']
            M = []
            for row in ws:
                line = []
                for cell in row:
                    line.append(cell.value)
                M.append(line)
            self.CommodityIntake = {}
            for r in range(8,len(M)):
                # Verify master data integrity
                com = M[r][1] # Commodity
                if com == None:
                    continue
                elif com not in self.MD_Commodities:
                    self.missing.append(["Commodity",com,"Commodity Intakes"])
                    continue
                # Load data
                self.CommodityIntake[com] = M[r][3]
            t_e = time.time()
            print "Duration: ", "{0:.3f}".format(t_e-t_s), " seconds"

        def shipping_leadtimes():
            print "> Loading Shipping Leadtimes"
            t_s = time.time()
            ws = self.db['3.04 Shipping Leadtimes']
            M = []
            for row in ws:
                line = []
                for cell in row:
                    line.append(cell.value)
                M.append(line)
            self.Shipping_Duration = {}
            for r in range(8,len(M)):
                # Verify master data integrity
                lp = M[r][2] # Load Port
                if lp == None:
                    continue
                elif lp not in self.MD_Locations_NDPs:
                    self.missing.append(["Named Delivery Place",lp,"Shipping Leadtimes"])
                    continue
                dp = M[r][4] # Discharge Port
                if dp not in self.MD_Locations_DPs:
                    self.missing.append(["Discharge Port",dp,"Shipping Leadtimes"])
                    continue
                # Load data
                dur0 = int(M[r][5]) # number of days
                rem = dur0 % 32 # remainder
                if rem < 15: # round down
                    dur = dur0 - rem + 1
                else: # round up
                    dur = dur0 - rem + 32
                self.Shipping_Duration[lp,dp] = dur
            t_e = time.time()
            print "Duration: ", "{0:.3f}".format(t_e-t_s), " seconds"

        def bulk_specifications():
            print "> Loading Shipping Specifications"
            t_s = time.time()
            ws = self.db['Shipping Specifications']
            M = []
            for row in ws:
                line = []
                for cell in row:
                    line.append(cell.value)
                M.append(line)
            self.Bulk_Tonnage = {}
            self.Bulk_Discount = {}
            for r in range(8,len(M)):
                # Verify master data integrity
                t = M[r][1] # Shipping type
                if t == None:
                    continue
                elif t not in self.MD_ShippingTypes:
                    self.missing.append(["Shipping Type",t,"Shipping Specifications"])
                    continue
                # Load data
                mt = M[r][2] # Metric Tonnes
                dc = M[r][3] # discount
                self.Bulk_Tonnage[t] = mt
                self.Bulk_Discount[t] = float(dc)
            t_e = time.time()
            print "Duration: ", "{0:.3f}".format(t_e-t_s), " seconds"

        def bulk_commodities():
            print "> Loading Bulk Commodities"
            t_s = time.time()
            ws = self.db['Bulk Commodities']
            M = []
            for row in ws:
                line = []
                for cell in row:
                    line.append(cell.value)
                M.append(line)
            self.Bulk_Commodity = {}
            for r in range(8,len(M)):
                # Verify master data integrity
                kt = M[r][1] # Commodity Type
                if kt == None:
                    continue
                elif kt not in self.MD_CommodityTypes:
                    self.missing.append(["Commodity Type",kt,"Bulk Commodities"])
                    continue
                t = M[r][2] # Shipping type
                if t not in self.MD_ShippingTypes:
                    self.missing.append(["Shipping Type",t,"Bulk Commodities"])
                    continue
                # Load data
                i = M[r][3] # Binary index
                if i == 1:
                    if t not in self.Bulk_Commodity.keys():
                        self.Bulk_Commodity[t] = []
                    self.Bulk_Commodity[t].append(kt)
            t_e = time.time()
            print "Duration: ", "{0:.3f}".format(t_e-t_s), " seconds"

        def bulk_lanes():
            print "> Loading Bulk Lanes"
            t_s = time.time()
            ws = self.db['Bulk Lanes']
            M = []
            for row in ws:
                line = []
                for cell in row:
                    line.append(cell.value)
                M.append(line)
            self.Bulk_Lanes = []
            for r in range(8,len(M)):
                # Verify master data integrity
                t = M[r][3] # Shipping type
                if t == None:
                    continue
                elif t not in self.MD_ShippingTypes:
                    self.missing.append(["Shipping Type",t,"Bulk Lanes"])
                    continue
                lp = M[r][1]
                if lp == None:
                    lp = ""
                elif lp not in self.MD_Locations_NDPs:
                    self.missing.append(["Load Port",lp,"Bulk Lanes"])
                    continue
                dp = M[r][2]
                if dp not in self.MD_Locations_DPs:
                    self.missing.append(["Discharge Port",dp,"Bulk Lanes"])
                    continue
                # Load data
                self.Bulk_Lanes.append((lp,dp,t))
            t_e = time.time()
            print "Duration: ", "{0:.3f}".format(t_e-t_s), " seconds"

        def landtransport_cost():
            print "> Loading Land Transport Costs"
            t_s = time.time()
            ws = self.db['3.06 Land Transport Cost']
            M = []
            for row in ws:
                line = []
                for cell in row:
                    line.append(cell.value)
                M.append(line)
            self.Transport_Cost = {}
            for r in range(8,len(M)):
                # Verify master data integrity
                orig = M[r][2] # Origin
                if orig == None:
                    continue
                elif orig not in self.MD_Locations:
                    self.missing.append(["Location",orig,"Land Transport Costs"])
                    continue
                dest = M[r][5] # Destination
                if dest not in self.MD_Locations:
                    self.missing.append(["Location",dest,"Land Transport Costs"])
                    continue
                cur = M[r][8]
                if cur not in self.MD_Currencies:
                    self.missing.append(["Currency",cur,"Land Transport Costs"])
                    continue
                elif cur not in self.CurrencyConversion.keys():
                    self.missing.append(["Currency Conversion",cur,"Land Transport Costs"])
                    continue
                val = M[r][7]
                if val == None:
                    continue
                # Load data
                self.Transport_Cost[orig,dest] = self.CurrencyConversion[cur] * val
            t_e = time.time()
            print "Duration: ", "{0:.3f}".format(t_e-t_s), " seconds"

        def landtransport_leadtimes():
            print "> Loading Land Transport Leadtimes"
            t_s = time.time()
            ws = self.db['3.07 Land Transport Leadtime']
            M = []
            for row in ws:
                line = []
                for cell in row:
                    line.append(cell.value)
                M.append(line)
            self.Transport_Dur = {}
            for r in range(8,len(M)):
                # Verify master data integrity
                orig = M[r][2] # Origin
                if orig == None:
                    continue
                elif orig not in self.MD_Locations:
                    self.missing.append(["Location",orig,"Land Transport Leadtimes"])
                    continue
                dest = M[r][5] # Destination
                if dest not in self.MD_Locations:
                    self.missing.append(["Location",dest,"Land Transport Leadtimes"])
                    continue
                # Load data
                dur0 = int(M[r][7]) # number of days
                rem = dur0 % 32 # remainder
                if rem < 15: # round down
                    dur = dur0 - rem + 1
                else: # round up
                    dur = dur0 - rem + 32
                self.Transport_Dur[orig,dest] = dur
            t_e = time.time()
            print "Duration: ", "{0:.3f}".format(t_e-t_s), " seconds"

        def dischargeports():
            print "> Loading Discharge Ports"
            t_s = time.time()
            ws = self.db['4.01 Discharge Ports']
            M = []
            for row in ws:
                line = []
                for cell in row:
                    line.append(cell.value)
                M.append(line)
            self.DischargePorts = {}
            for r in range(8,len(M)):
                # Verify master data integrity
                c = M[r][1] # Country
                if c == None:
                    continue
                elif c not in self.MD_Countries:
                    self.missing.append(["Country",c,"Discharge Ports"])
                    continue
                dp = M[r][2] # Discharge Port
                if dp not in self.MD_Locations_DPs:
                    self.missing.append(["Discharge Port",dp,"Discharge Ports"])
                    continue
                # Load data
                if c not in self.DischargePorts.keys():
                    self.DischargePorts[c] = []
                self.DischargePorts[c].append(dp)
            t_e = time.time()
            print "Duration: ", "{0:.3f}".format(t_e-t_s), " seconds"

        def port_cost():
            print "> Loading Port Handling Costs"
            t_s = time.time()
            ws = self.db['4.02 Port Handling Cost']
            M = []
            for row in ws:
                line = []
                for cell in row:
                    line.append(cell.value)
                M.append(line)
            self.Handling_Cost = {}
            for r in range(8,len(M)):
                # Verify master data integrity
                dp = M[r][1] # Discharge Port
                if dp == None:
                    continue
                elif dp not in self.MD_Locations_DPs:
                    self.missing.append(["Discharge Port",dp,"Port Handling Costs"])
                    continue
                cur = M[r][4]
                if cur not in self.MD_Currencies:
                    self.missing.append(["Currency",cur,"Port Handling Costs"])
                    continue
                elif cur not in self.CurrencyConversion.keys():
                    self.missing.append(["Currency Conversion",cur,"Port Handling Costs"])
                    continue
                t = M[r][2]
                if t not in self.MD_ShippingTypes:
                    self.missing.append(["Shipping Type",t,"Port Handling Costs"])
                    continue
                val = M[r][3]
                if val == None:
                    continue
                # Load data
                self.Handling_Cost[dp,t] = self.CurrencyConversion[cur] * val
            t_e = time.time()
            print "Duration: ", "{0:.3f}".format(t_e-t_s), " seconds"

        def port_capacity():
            print "> Loading Port Handling Capacity"
            t_s = time.time()
            ws = self.db['4.03 Port Handling Capacity']
            M = []
            for row in ws:
                line = []
                for cell in row:
                    line.append(cell.value)
                M.append(line)
            self.Handling_Capacity = {}
            for r in range(8,len(M)):
                # Verify master data integrity
                dp = M[r][1] # Discharge Port
                if dp == None:
                    continue
                elif dp not in self.MD_Locations_DPs:
                    self.missing.append(["Discharge Port",dp,"Port Handling Capacity"])
                    continue
                t = M[r][2] # type
                if t not in self.MD_ShippingTypes:
                    self.missing.append(["Shipping Type",t,"Port Handling Capacity"])
                    continue
                d = M[r][4] # date
                # Load data
                self.Handling_Capacity[dp,t,d] = M[r][3]
            t_e = time.time()
            print "Duration: ", "{0:.3f}".format(t_e-t_s), " seconds"

        def port_storage():
            print "> Loading Port Storage"
            t_s = time.time()
            ws = self.db['4.04 Port Storage']
            M = []
            for row in ws:
                line = []
                for cell in row:
                    line.append(cell.value)
                M.append(line)
            self.Storage_Cost = {}
            self.Storage_Capacity = {}
            for r in range(8,len(M)):
                # Verify master data integrity
                dp = M[r][1] # Discharge Port
                if dp == None:
                    continue
                elif dp not in self.MD_Locations_DPs:
                    self.missing.append(["Discharge Port",dp])
                    continue
                cur = M[r][6]
                if cur not in self.MD_Currencies:
                    self.missing.append(["Currency",cur])
                    continue
                elif cur not in self.CurrencyConversion.keys():
                    self.missing.append(["Currency Conversion",cur])
                    continue
                # Load data
                mt = M[r][2]
                if mt != None:
                    self.Storage_Capacity[dp] = mt
                val = M[r][5]
                if val != None:
                    self.Storage_Cost[dp] = self.CurrencyConversion[cur] * val
            t_e = time.time()
            print "Duration: ", "{0:.3f}".format(t_e-t_s), " seconds"

        def port_leadtimes():
            print "> Loading Port Leadtimes"
            t_s = time.time()
            ws = self.db['4.05 Port Leadtimes']
            M = []
            for row in ws:
                line = []
                for cell in row:
                    line.append(cell.value)
                M.append(line)
            self.Handling_Dur = {}
            for r in range(8,len(M)):
                # Verify master data integrity
                dp = M[r][3] # Discharge Port
                if dp == None:
                    continue
                elif dp not in self.MD_Locations_DPs:
                    self.missing.append(["Discharge Port",dp])
                    continue
                # Load data
                dur0 = int(M[r][7]) # number of days
                rem = dur0 % 32 # remainder
                if rem < 15: # round down
                    dur = dur0 - rem
                else: # round up
                    dur = dur0 - rem + 32
                self.Handling_Dur[dp] = dur
            t_e = time.time()
            print "Duration: ", "{0:.3f}".format(t_e-t_s), " seconds"

        def edp_storage():
            print "> Loading EDP Storage"
            t_s = time.time()
            ws = self.db['4.09 EDP Storage']
            M = []
            for row in ws:
                line = []
                for cell in row:
                    line.append(cell.value)
                M.append(line)
            self.Storage_Capacity_Country = {}
            for r in range(8,len(M)):
                # Verify master data integrity
                edp = M[r][2] # Extended Delivery Point
                if edp == None:
                    continue
                elif edp not in self.MD_Locations_EDPs:
                    self.missing.append(["Extended Delivery Point",edp,"EDP Storage"])
                    continue
                cur = M[r][7]
                if cur not in self.MD_Currencies:
                    self.missing.append(["Currency",cur,"EDP Storage"])
                    continue
                elif cur not in self.CurrencyConversion.keys():
                    self.missing.append(["Currency Conversion",cur,"EDP Storage"])
                    continue
                # Load data
                val = M[r][6]
                if val != None:
                    self.Storage_Cost[edp] = self.CurrencyConversion[cur] * val
                mt = M[r][3]
                if mt != None and mt > 0:
                    self.Storage_Capacity[edp] = mt
                    c = M[r][1]
                    if c in self.Storage_Capacity_Country.keys():
                        self.Storage_Capacity_Country[c] += mt
                    else:
                        self.Storage_Capacity_Country[c] = mt
            t_e = time.time()
            print "Duration: ", "{0:.3f}".format(t_e-t_s), " seconds"

        # Execute
        t_0 = time.time()
        print "Loading Logistics Data"
        container_rates()
        commodity_intakes()
        shipping_leadtimes()
        bulk_specifications()
        bulk_commodities()
        bulk_lanes()
        landtransport_cost()
        landtransport_leadtimes()
        dischargeports()
        port_cost()
        port_capacity()
        port_storage()
        port_leadtimes()
        edp_storage()
        t_e = time.time()
        print "Finished loading logistics data in ", "{0:.3f}".format(t_e-t_0), " seconds"
        print " "

    def load_data_resources(self):
        '''
        Unfiltered loading of the resource data
        '''

        def donation():
            print "> Loading In-kind Donations"
            t_s = time.time()
            ws = self.db['5.02 IK Donations']
            M = []
            for row in ws:
                line = []
                for cell in row:
                    line.append(cell.value)
                M.append(line)
            self.Donation = {}
            for r in range(8,len(M)):
                # Verify master data integrity
                sk = M[r][2] # commodity
                if sk == None:
                    continue
                elif sk not in self.MD_Commodities:
                    self.missing.append(["Commodity",sk,"In-Kind Donations"])
                    continue
                don = M[r][1] # donor
                qty = M[r][3] # quantity
                o = M[r][4] # origin
                if o not in self.MD_Locations:
                    self.missing.append(["Location",o,"In-Kind Donations"])
                    continue
                aod = M[r][5] # as of date
                dur = M[r][6] # duration (months)
                cor = M[r][7] # corridor (optional)
                if cor not in self.MD_Locations and cor != None:
                    self.missing.append(["Location",cor,"In-Kind Donations"])
                    continue
                c = M[r][8]
                if c not in self.MD_Countries:
                    self.missing.append(["Country",c,"In-Kind Donations"])
                    continue
                p = str(M[r][9]) # Project
                if p not in self.MD_Projects:
                    self.missing.append(["Project",p,"In-Kind Donations"])
                    continue
                # Load info
                self.Donation[don,sk,o,aod,dur,cor,c,p] = qty
            t_e = time.time()
            print "Duration: ", "{0:.3f}".format(t_e-t_s), " seconds"

        def unprogrammed_funds():
            print "> Loading Unprogrammed Funds"
            t_s = time.time()
            ws = self.db['5.01 Funding Availability']
            M = []
            for row in ws:
                line = []
                for cell in row:
                    line.append(cell.value)
                M.append(line)
            self.Unprogrammed = []
            for r in range(8,len(M)):
                # Verify master data integrity
                rc = M[r][1] # Recipient Country
                if rc == None:
                    continue
                elif rc not in self.CONV_Country.keys():
                    self.missing.append(["Country Conversion",rc,"Unprogrammed Funds"])
                    continue
                else:
                    rc = self.CONV_Country[rc]
                p = str(M[r][2]) # Project
                if p not in self.MD_Projects:
                    self.missing.append(["Project",p,"Unprogrammed Funds"])
                    continue
                tm = M[r][3] # Transfer Modality
                if M[r][4]!="#VALUE!":
                    cod = M[r][4] # Created On Date
                else:
                    continue
                if M[r][5]!="#VALUE!":
                    vfd = M[r][5] # Valid From Date
                else:
                    continue
                usd = M[r][6] # Unprogrammed Funds
                # Load info
                d = cod if diff_month(cod,vfd)>0 else vfd
                self.Unprogrammed.append([rc,p,tm,d,usd])
            t_e = time.time()
            print "Duration: ", "{0:.3f}".format(t_e-t_s), " seconds"

        def ltsh():
            print "> Loading LTSH Rates"
            t_s = time.time()
            ws = self.db['5.04 LTSH Rates']
            M = []
            for row in ws:
                line = []
                for cell in row:
                    line.append(cell.value)
                M.append(line)
            self.LTSH = {}
            for r in range(8,len(M)):
                # Verify master data integrity
                c = M[r][1] # country
                if c == None:
                    continue
                elif c not in self.MD_Countries:
                    if c in self.CONV_Country.keys():
                        c = self.CONV_Country[c]
                    else:
                        self.missing.append(["Country",c,"LTSH"])
                        continue
                p = str(M[r][2]) # Project
                if p not in self.MD_Projects:
                    self.missing.append(["Project",p,"LTSH"])
                    continue
                # Load info
                self.LTSH[c,p] = {}
                self.LTSH[c,p]["Local"] = (M[r][3],M[r][4]) # (OVL, ITSH)
                self.LTSH[c,p]["Regional"] = (M[r][5],M[r][6])
                self.LTSH[c,p]["Overseas"] = (M[r][7],M[r][8])
                self.LTSH[c,p]["Average"] = (M[r][9],M[r][10])

            t_e = time.time()
            print "Duration: ", "{0:.3f}".format(t_e-t_s), " seconds"

        def odoc():
            print "> Loading ODOC Rates"
            t_s = time.time()
            ws = self.db['5.05 ODOC Rates']
            M = []
            for row in ws:
                line = []
                for cell in row:
                    line.append(cell.value)
                M.append(line)
            self.ODOC = {}
            for r in range(8,len(M)):
                # Verify master data integrity
                c = M[r][1] # country
                if M[r][4] == None:
                    continue
                elif c not in self.MD_Countries:
                    self.missing.append(["Country",c,"ODOC"])
                    continue
                p = str(M[r][2]) # Project
                if p not in self.MD_Projects:
                    self.missing.append(["Project",p,"ODOC"])
                    continue
                # Load info
                self.ODOC[c,p] = M[r][4]
            t_e = time.time()
            print "Duration: ", "{0:.3f}".format(t_e-t_s), " seconds"

        def dsc():
            print "> Loading DSC Rates"
            t_s = time.time()
            ws = self.db['5.07 DSC Rates']
            M = []
            for row in ws:
                line = []
                for cell in row:
                    line.append(cell.value)
                M.append(line)
            self.DSC = {}
            for r in range(8,len(M)):
                # Verify master data integrity
                c = M[r][1] # country
                if M[r][4] == None:
                    continue
                elif c not in self.MD_Countries:
                    self.missing.append(["Country",c,"DSC"])
                    continue
                p = str(M[r][2]) # Project
                if p not in self.MD_Projects:
                    self.missing.append(["Project",p,"DSC"])
                    continue
                # Load info
                self.DSC[c,p] = M[r][4]
            t_e = time.time()
            print "Duration: ", "{0:.3f}".format(t_e-t_s), " seconds"

        def fcr():
            print "> Loading FCR Rates"
            t_s = time.time()
            ws = self.db['FCR Rates (By Com)']
            M = []
            for row in ws:
                line = []
                for cell in row:
                    line.append(cell.value)
                M.append(line)
            self.FCR = {}
            for r in range(2,len(M)):
                # Verify master data integrity
                c = M[r][1].upper() # Country
                if c == None:
                    continue
                elif c not in self.MD_Countries:
                    if c in self.CONV_Country.keys():
                        c = self.CONV_Country[c]
                    else:
                        self.missing.append(["Country",c,"FCR Rates"])
                        continue
                p = str(M[r][2]) # Project
                if p not in self.MD_Projects:
                    self.missing.append(["Project",p,"FCR Rates"])
                    continue
                k0 = M[r][3].upper() # Commodity
                if (p,k0) in self.CONV_Commodity.keys():
                    k = self.CONV_Commodity[p,k0]
                elif ('*',k0) in self.CONV_Commodity.keys():
                    k = self.CONV_Commodity['*',k0]
                else:
                    self.missing.append(["Commodity Conversion",k0 + " @ " + c + "_" + p,"FCR Rates"])
                    continue
                fcr = M[r][4] # FCR Rate
                # Load info
                if fcr != None:
                    self.FCR[c,p,k] = fcr

            ws = self.db['FCR Rates (By Project)']
            M = []
            for row in ws:
                line = []
                for cell in row:
                    line.append(cell.value)
                M.append(line)
            for r in range(2,len(M)):
                # Verify master data integrity
                c = M[r][1].upper() # Country
                if c == None:
                    continue
                elif c not in self.MD_Countries:
                    if c in self.CONV_Country.keys():
                        c = self.CONV_Country[c]
                    else:
                        self.missing.append(["Country",c,"FCR Rates"])
                        continue
                p = str(M[r][2]) # Project
                if p not in self.MD_Projects:
                    self.missing.append(["Project",p,"FCR Rates"])
                    continue
                fcr = M[r][5] # FCR Rate
                # Load info
                if fcr != None:
                    self.FCR[c,p] = fcr

            t_e = time.time()
            print "Duration: ", "{0:.3f}".format(t_e-t_s), " seconds"

        def currency_conversion():
            print "> Loading Currency Conversion"
            t_s = time.time()
            ws = self.db['5.11 Currency Conversion']
            M = []
            for row in ws:
                line = []
                for cell in row:
                    line.append(cell.value)
                M.append(line)
            self.CurrencyConversion = {}
            for r in range(8,len(M)):
                # Verify master data integrity
                cur = M[r][1]
                if cur == None:
                    continue
                elif cur not in self.MD_Currencies:
                    self.missing.append(["Currency",cur,"Currency Conversion"])
                    continue
                d = M[r][3]
                # Load info
                self.CurrencyConversion[cur] = float(M[r][2])
            t_e = time.time()
            print "Duration: ", "{0:.3f}".format(t_e-t_s), " seconds"

        # Execute
        t_0 = time.time()
        print "Loading Resource Data"
        donation()
        unprogrammed_funds()
        ltsh()
        odoc()
        dsc()
        fcr()
        currency_conversion()
        t_e = time.time()
        print "Finished loading resources data in ", "{0:.3f}".format(t_e-t_0), " seconds"
        print " "

    def load_data_transactional(self):
        '''
        Unfiltered loading of transactional data
        '''

        def inventory():
            print "> Loading Inventory"
            t_s = time.time()
            ws = self.db['5.10 Inventory']
            M = []
            for row in ws:
                line = []
                for cell in row:
                    line.append(cell.value)
                M.append(line)
            self.Inventory = {} # (loc,k,c,p) -> [inventory, in transit]
            self.GCMF = {}
            for r in range(8,len(M)):
                # Verify master data integrity
                rb = M[r][1] # Regional Bureau
                rc = M[r][2] # Recipient Country
                wbs = M[r][3] # WBS element
                if wbs == None:
                    continue
                p = wbs[:6] # Project
                sc = M[r][4] # Storage Location Country
                if sc not in self.MD_Countries:
                    self.missing.append(["Country",sc,"Inventory"])
                    continue
                k0 = M[r][5].upper() # Commodity (LESS)
                if k0.startswith("UHC"):
                    continue
                elif (p,k0) in self.CONV_Commodity.keys():
                    k = self.CONV_Commodity[p,k0]
                elif ('*',k0) in self.CONV_Commodity.keys():
                    k = self.CONV_Commodity['*',k0]
                else:
                    self.missing.append(["Commodity Conversion",k0 + " @ " + rc + "_" + p,"Inventory"])
                    continue
                if wbs[:2] != "S.": # CO owned inventory
                    if rc == "null":
                        try:
                            rcs = self.MD_Projects_Countries[p] # find all RCs that are linked to the project
                            if len(rcs) == 1: # if there's only one, we know what to do
                                rc = rcs[1]
                        except:
                            self.missing.append(["Recipient Country","null","Inventory"])
                            continue
                    if rc not in self.MD_Countries:
                        self.missing.append(["Recipient Country",rc,"Inventory"])
                        continue
                    if (rc,sc) in self.CONV_LESS_Location.keys():
                        loc = self.CONV_LESS_Location[rc,sc] # loc is a CO, NDP, or DP
                    elif rc == sc:
                        loc = rc # loc is a CO
                    else:
                        self.missing.append(["Location Mapping (LESS)",rc + ", " + sc,"Inventory"])
                        continue
                    inv = sum(M[r][i] for i in [11,12] if M[r][i] != None) # Inventory (MT)
                    itr = sum(M[r][i] for i in [9,10,13] if M[r][i] != None) # In Transit (MT)
                    # Load info
                    key = (loc,k,rc,p)
                    if key not in self.Inventory.keys():
                        self.Inventory[key] = [0,0]
                    if inv != None:
                        self.Inventory[key][0] += inv
                    if itr != None:
                        self.Inventory[key][1] += itr
                else: # GCMF inventory
                    if sc in self.CONV_LESS_LocationLink.keys():
                        loc = self.CONV_LESS_LocationLink[sc]
                    else:
                        self.missing.append(["Location Mapping (LESS)","GCMF @ " + sc,"Inventory"])
                        continue
                    inv = sum(M[r][i] for i in [7,8] if M[r][i] != None) # Inventory (MT)
                    itr = M[r][6] # In Transit (MT)
                    # Load info
                    key = (rb,loc,k)
                    if key not in self.GCMF.keys():
                        self.GCMF[key] = [0,0]
                    if inv != None:
                        self.GCMF[key][0] += inv
                    if itr != None:
                        self.GCMF[key][1] += itr

            t_e = time.time()
            print "Duration: ", "{0:.3f}".format(t_e-t_s), " seconds"

        def po():
            print "> Loading PO Report"
            t_s = time.time()
            ws = self.db['PO Report']
            M = []
            for row in ws:
                line = []
                for cell in row:
                    line.append(cell.value)
                M.append(line)
            self.OpenPOs = {} # POs that have to be entered into Quintiq through feedback
            self.GCMF_Price = {} # Latest price for GCMF stocks
            self.IK_Price = {} # Latest price for IK donations
            opts = {}
            po2scips = []
            ik2scips = []
            po_new = []
            zfpf = []
            history = []
            for r in range(8,len(M)):
                # Verify master data integrity
                poi = M[r][1] # PO Item (unique)
                pon = poi[:10] # PO Number (not unique)
                pod = M[r][2] # PO Date
                wbs = str(M[r][19]) # WBS
                p = wbs[:6] # Project
                if p not in self.MD_Projects:
                    self.missing.append(["Project",p,"PO Report"])
                    continue
                rc = M[r][12] # Recipient Country
                if rc in self.CONV_Country.keys():
                    rc = self.CONV_Country[rc]
                else:
                    if p.startswith("S"):
                        rc = "GCMF"
                    else:
                        self.missing.append(["RC Conversion",rc,"PO Report"])
                        continue
                k0 = M[r][8] # Procured Commodity
                if k0 == None or k0 == "":
                    continue
                k0 = k0.upper()
                if (p,k0) in self.CONV_Commodity.keys():
                    sk = self.CONV_Commodity[p,k0]
                elif ('*',k0) in self.CONV_Commodity.keys():
                    sk = self.CONV_Commodity['*',k0]
                elif k0 in self.MD_Commodities:
                    sk = k0
                else:
                    self.missing.append(["Commodity Conversion",k0 + " @ " + rc + "_" + p,"PO Report"])
                    continue
                usd = M[r][11] # $/MT
                mt = M[r][10] # MT purchased
                inc = M[r][15] # Incoterm
                b = M[r][20] # PO-GR Balance (MT)
                pot = M[r][5] # PO Type
                if pot == "POFC" or pot == "POFW": # int'l/reg'l/local procurement (incl. procurement by GCMF)
                    oc = M[r][9] # Origin Country
                    if oc not in self.MD_Countries:
                        self.missing.append(["Country",oc,"PO Report"])
                        continue
                    sd_f = M[r][13] # Ship From Date
                    sd_t = M[r][14] # Ship To Date
                    ndp = M[r][16] # Named Delivery Place
                    if ndp in self.CONV_NDP.keys():
                        ndp = self.CONV_NDP[ndp]
                        if ndp == 0: # Unmappable NDP
                            continue
                        if ndp not in self.MD_Locations_NDPs and ndp not in self.MD_Locations_DPs: # Mappable but forgot to add to MD
                            self.missing.append(["NDP",ndp,"PO Report"])
                            continue
                    else:
##                        if ndp in opts.keys():
##                            continue
##                        opts[ndp] = difflib.get_close_matches(ndp,self.MD_Locations_NDPs+self.MD_Locations_DPs,1,0.4)
##                        try:
##                            print ndp," | ",opts[ndp][0]
##                        except:
##                            print ""
                        self.missing.append(["NDP Conversion",ndp,"PO Report"])
                        continue
                    # Load info
                    key = (oc,ndp,sk,0,pod) # the 0 is the GMO indicator -> commodity assumed to be Non-GMO
                    if key in self.Procurement_Cost.keys():
                        if self.Procurement_Cost[key] < usd: # use the cheapest option if key is equal
                            continue
                    self.Procurement_Inco[key] = inc
                    self.Procurement_Cost[key] = usd
                    self.Procurement_Date[key] = pod # As Of Date not known/relevant
                    if b > 0: # Still waiting for some MTs -> use as feedback (open PO)
                        if p.startswith("S"): # GCMF -> need WBS
                            self.OpenPOs[rc,wbs,poi,sk,0,oc,ndp,sd_t] = b
                        else: # Normal procurement -> project is enough
                            self.OpenPOs[rc,p,poi,sk,0,oc,ndp,sd_t] = b
                    kt = self.MD_Commodities_Type[sk]
                    g = self.MD_Commodities_Group[sk]
                    skt = kt + " (FOOD - " + g + ")"
                    po2scips.append((rc,p,oc,inc,ndp,skt,sk,"Non GMO","Bag",usd,pod,100000,60,"PO"))
                    if pon not in self.PO_Origins:
                        self.PO_Origins[pon] = (oc,ndp)
                        po_new.append((pon,oc,ndp))
                    ic = self.MD_Locations_Country[ndp]
                    history.append((rc,p,sk,pot,oc,ic,ndp,mt))
                elif pot == "POFI": # Type = POFI -> In-Kind donation
                    ndp = M[r][16] # Named Delivery Place
                    if ndp in self.CONV_NDP.keys():
                        ndp = self.CONV_NDP[ndp]
                        if ndp == 0: # Unmappable NDP
                            continue
                        if ndp not in self.MD_Locations_NDPs and ndp not in self.MD_Locations_DPs: # Mappable but forgot to add to MD
                            self.missing.append(["NDP",ndp,"PO Report"])
                            continue
                    else:
                        self.missing.append(["NDP Conversion",ndp,"PO Report"])
                        continue
                    ic = self.MD_Locations_Country[ndp] # Incoterm country
                    oc = M[r][9] # Donor Country
                    if oc not in self.MD_Countries:
                        oc = ic
                    sd_t = M[r][14] # Ship To Date
                    dp0 = M[r][17]
                    if dp0 in self.CONV_DP.keys():
                        dp = self.CONV_DP[dp0]
                    else:
                        dp = None
                        self.missing.append(["DP Conversion",dp,"PO Report"])
                    # Load info
                    key = (oc,sk,ndp,sd_t,1,dp,rc,p)
                    if b >= 0:
                        self.Donation[key] = b
                    if (oc,ndp,sk) not in self.IK_Price.keys():
                        self.IK_Price[oc,ndp,sk] = (usd, pod)
                    elif pod > self.IK_Price[oc,ndp,sk][1]:
                        self.IK_Price[oc,ndp,sk] = (usd, pod)
                    kt = self.MD_Commodities_Type[sk]
                    g = self.MD_Commodities_Group[sk]
                    skt = kt + " (FOOD - " + g + ")"
                    ik2scips.append((rc,p,oc,inc,ndp,skt,sk,"Non GMO","Bag",usd,pod,100000,60,"POFI"))
                    history.append((rc,p,sk,pot,oc,ic,ndp,mt))
                else: # Type = ZFPF -> purchase from GCMF
                    # Load info
                    if (p,sk) not in self.GCMF_Price.keys():
                        self.GCMF_Price[p,sk] = (usd, pod)
                    elif pod > self.GCMF_Price[p,sk][1]: # This entry is newer
                        self.GCMF_Price[p,sk] = (usd, pod)
                    po = M[r][18][:10] # PO number linked to IPO
                    ipo = pon # IPO number
                    if po == None:
                        if ipo not in self.CONV_ZFPF.keys():
                            self.missing.append(["IPO Conversion",ipo,"PO Report"])
                            continue
                        po = self.CONV_ZFPF[ipo]
                    if po not in self.PO_Origins.keys():
                        self.missing.append(["PO Origin",po,"PO Report"])
                        continue
                    oc,ndp = self.PO_Origins[po]
                    zfpf.append((rc,p,ipo,po,pod,oc,ndp,sk,usd,mt))
                    ic = self.MD_Locations_Country[ndp]
                    history.append((rc,p,sk,pot,oc,ic,ndp,mt))

            print "Exporting PO conversions to 'Debugging.xlsx'"
            header = ["Recipient Country","Project","Origin Country","Incoterm","NDP",
                "Commodity Type","Specific Commodity","GMO","Packaging","Price ($/MT)","PO Date",
                "Capacity (MT/month)","Lead Time (PR to GR)","Source"]
            self.print_to_file(po2scips,"SCIPS POs",header)
            self.print_to_file(ik2scips,"SCIPS POFIs",header)
            header = ["PO Number","Origin Country","Named Delivery Place"]
            self.print_to_file(po_new,"Unmapped POs",header)
            header = ["Recipient Country","Project","IPO","PO","PO Date","Origin Country","NDP",
                "Specific Commodity","Price ($/MT)","Quantity (MT)"]
            self.print_to_file(zfpf,"ZFPF Mapping",header)
            header = ["Recipient Country","Project","Specific Commodity",
                "PO Type","Origin Country","Incoterm Country","NDP","Quantity (MT)"]
            self.print_to_file(history,"PO longlist",header)
            shortlist = {}
            ag = {}
            ag_country = {}
            key_list = {}
            for rc,p,sk,pot,oc,ic,ndp,mt in history:
                if mt == 0:
                    continue
                key = (rc,sk)
                if key not in key_list.keys():
                    key_list[key] = []
                    shortlist[key] = {}
                    ag[key] = 0
                if ic not in key_list[key]:
                    key_list[key].append(ic)
                    ag_country[rc,sk,ic] = 0
                ag[key] += mt
                ag_country[rc,sk,ic] += mt
            for rc,sk in ag.keys():
                tot = ag[rc,sk]
                for ic in key_list[rc,sk]:
                    mt = ag_country[rc,sk,ic]
                    if mt >= 0.1 * tot:
                        shortlist[rc,sk][ic] = mt
                    else:
                        if "Other" not in shortlist[rc,sk].keys():
                            shortlist[rc,sk]["Other"] = 0
                        shortlist[rc,sk]["Other"] += mt
            temp = []
            for rc,sk in shortlist.keys():
                for ic in shortlist[rc,sk].keys():
                    mt = shortlist[rc,sk][ic]
                    temp.append([rc,sk,ic,mt])
            header = ["Recipient Country","Specific Commodity","Incoterm Country","Quantity (MT)"]
            self.print_to_file(temp,"PO shortlist",header)
            self.allocation = {}
            for rc,sk in shortlist.keys():
                tot = ag[rc,sk]
                l = len(shortlist[rc,sk])
                if "Other" in shortlist[rc,sk].keys():
                    bonus = shortlist[rc,sk]["Other"] / (l-1)
                else:
                    bonus = 0
                for ic in shortlist[rc,sk]:
                    if ic == "Other":
                        continue
                    mt = shortlist[rc,sk][ic] + bonus
                    self.allocation[rc,sk,ic] = mt/tot
            header = ["Recipient Country","Specific Commodity","Incoterm Country","Allocation (%)"]
            self.print_to_file(self.allocation,"Sourcing Allocation",header)

            t_e = time.time()
            print "Duration: ", "{0:.3f}".format(t_e-t_s), " seconds"

        # Execute
        t_0 = time.time()
        print "Loading Transactional Data"
        inventory()
        po()
        t_e = time.time()
        print "Finished loading Transactional Data in ", "{0:.3f}".format(t_e-t_0), " seconds"
        print " "

    def load_data_other(self):
        '''
        Unfiltered loading of the resource data
        '''

        def project_priorities():
            print "> Loading Project Priorities"
            t_s = time.time()
            ws = self.db['Project Priorities']
            M = []
            for row in ws:
                line = []
                for cell in row:
                    line.append(cell.value)
                M.append(line)
            self.Project_Priority = {}
            for r in range(8,len(M)):
                # Verify master data integrity
                c = M[r][2]
                if c not in self.MD_Countries:
                    self.missing.append(["Country",c,"Project Priorities"])
                    continue
                p = str(M[r][4]) # project
                if p == None:
                    continue
                elif p not in self.MD_Projects:
                    self.missing.append(["Project",p,"Project Priorities"])
                    continue
                # Load info
                self.Project_Priority[c,p] = M[r][5]
            t_e = time.time()
            print "Duration: ", "{0:.3f}".format(t_e-t_s), " seconds"

        def commodity_priorities():
            print "> Loading Commodity Priorities"
            t_s = time.time()
            ws = self.db['Commodity Priorities']
            M = []
            for row in ws:
                line = []
                for cell in row:
                    line.append(cell.value)
                M.append(line)
            self.Commodity_Priority = {}
            for r in range(8,len(M)):
                # Verify master data integrity
                p = str(M[r][3]) # project
                if p == None:
                    continue
                elif p not in self.MD_Projects:
                    self.missing.append(["Project",p,"Commodity Priorities"])
                    continue
                com = M[r][4] # commodity (tact demand)
                if (p,com) in self.CONV_Commodity.keys():
                    sk = self.CONV_Commodity[p,com] # specific commodity
                elif ("*",com) in self.CONV_Commodity.keys():
                    sk= self.CONV_Commodity["*",com]
                elif com in self.MD_Commodities:
                    sk = com
                else:
                    self.missing.append(["Commodity Conversion",com + "_" + p,"Commodity Priorities"])
                    continue
                # Load info
                self.Commodity_Priority[sk,p] = M[r][5]
            t_e = time.time()
            print "Duration: ", "{0:.3f}".format(t_e-t_s), " seconds"

        def demand_selection():
            print "> Loading Demand Selection"
            t_s = time.time()
            ws = self.db['Demand Selection']
            M = []
            for row in ws:
                line = []
                for cell in row:
                    line.append(cell.value)
                M.append(line)
            self.Demand_Selection = {}
            for r in range(2,len(M)):
                # Verify master data integrity
                c = M[r][1]
                if c == None:
                    continue
                elif c not in self.MD_Countries:
                    self.missing.append(["Country",c,"Demand Selection"])
                    continue
                # Load info
                self.Demand_Selection[c] = (M[r][2],M[r][3])
            t_e = time.time()
            print "Duration: ", "{0:.3f}".format(t_e-t_s), " seconds"

        # Execute
        t_0 = time.time()
        print "Loading Other Data"
        project_priorities()
        commodity_priorities()
        demand_selection()
        t_e = time.time()
        print "Finished loading resources data in ", "{0:.3f}".format(t_e-t_0), " seconds"
        print " "

    def load_data_constraints(self):
        '''
        Unfiltered loading of constraints (user inputs)
        '''

        def sourcing_restrictions():
            print "> Loading Sourcing Restrictions"
            t_s = time.time()
            ws = self.db['6.04 Sourcing Restrictions']
            M = []
            for row in ws:
                line = []
                for cell in row:
                    line.append(cell.value)
                M.append(line)
            self.Sourcing_Restrictions = {}
            for r in range(8,len(M)):
                # Verify master data integrity
                oc = M[r][1].upper() # origin country
                if oc == None:
                    continue
                elif oc not in self.MD_Countries:
                    self.missing.append(["Country",oc,"Sourcing Restrictions"])
                    continue
                k = M[r][2] # specific commodity or commodity type
                if k not in self.MD_Commodities and k not in self.MD_CommodityTypes:
                    self.missing.append(["Commodity",k,"Sourcing Restrictions"])
                    continue
                c = M[r][3] # country (optional)
                if c != None and c not in self.MD_Countries:
                    self.missing.append(["Country",c,"Sourcing Restrictions"])
                    continue
                p = M[r][4] # project (optional)
                if p != None:
                    p = str(p)
                    if p not in self.MD_Projects:
                        self.missing.append(["Project",p,"Sourcing Restrictions"])
                        continue
                d0 = M[r][5] # start date
                d1 = M[r][6] # end date (optional)
                ag = M[r][7] # aggregate (binary)
                cap = M[r][8] # capacity
                # Load info
                self.Sourcing_Restrictions[oc,k,c,p,d0,d1,ag] = cap
            t_e = time.time()
            print "Duration: ", "{0:.3f}".format(t_e-t_s), " seconds"

        # Execute
        t_0 = time.time()
        print "Loading Constraints"
        sourcing_restrictions()
        t_e = time.time()
        print "Finished loading Constraints in ", "{0:.3f}".format(t_e-t_0), " seconds"
        print " "

    def print_missing(self):
        '''
        Notify the user of any data mismatches
        '''
        self.missing = [list(t) for t in set(tuple(element) for element in self.missing)]
        n = len(self.missing)
        if n > 0:
            path0 = os.path.join(self.data_dir,"Missing.xlsx")
            path1 = os.path.join(self.script_dir,"Missing.xlsx")
            wb = xl.load_workbook(filename = path0, read_only=False)
            ws = wb.active
            print "<<<WARNING>>>"
            print "Unable to match to Master Data (" + str(n) + " cases)"
            if n <= 30:
                for i in sorted(self.missing):
                    print i
                    ws.append([i[0],i[1],i[2]])
            else:
                print "More than 30 issues detected -> output suppressed"
                for i in sorted(self.missing):
                    ws.append([i[0],i[1],i[2]])
            print "All cases have been stored in an Excel file at:"
            print path1
            print "<<<WARNING>>>"
            print " "
            wb.save(filename = path1)

    def print_to_file(self,LIST,NAME,HEADER):
        f = os.path.join(self.script_dir, "Debugging.xlsx")
        db = xl.load_workbook(filename = f, read_only=False, data_only=True)
        if NAME in db.sheetnames:
            sheet = db.get_sheet_by_name(NAME)
            db.remove_sheet(sheet)
        db.create_sheet(NAME)
        ws = db.get_sheet_by_name(NAME)
        ws.append(HEADER)
        try:
            check = LIST.keys()
            check = 1
        except:
            check = 0
        if check == 1: # LIST = dictionary
            for key in sorted(LIST.keys()):
                # turn pointer into tuple
                if isinstance(key,tuple):
                    t_key = key
                else:
                    t_key = (key,)
                # turn value into tuple
                v = LIST[key]
                if isinstance(v,tuple):
                    row = t_key + v
                elif isinstance(v,list):
                    row = t_key + tuple(v)
                else: # single value
                    row = t_key + (v,)
                # print row
                try:
                    ws.append(row)
                except:
                    row = tuple(j for i in row for j in (i if isinstance(i, tuple) else (i,))) # untangle tuples
                    ws.append(row)
        else: # LIST = list
            LIST = [list(t) for t in set(tuple(element) for element in LIST)]
            LIST.sort()
            for row in LIST:
                ws.append(row)
        db.save(filename = f)

    def remove_bs(self):
        '''
        Filter data based on selected RB
        '''
        def projects():
            print "> Establishing relevant projects"
            t_s = time.time()
            self.Q_SalesSegments = [] # Country-Project combinations
            self.Q_Projects = [] # subset of Projects
            self.Q_RBs = [] # subset of RBs
            if self.Q_COs == []: # self.Q_RB defines selection of salessegments
                for p in self.MD_Projects:
                    if self.MD_Projects_RB[p] == self.Q_RB:
                        self.Q_Projects.append(p)
                        for c in self.MD_Projects_Countries[p]:
                            self.Q_SalesSegments.append((c,p))
                            self.Q_COs.append(c)
                self.Q_COs = list(set(self.Q_COs))
                self.Q_COs.sort()
            else:
                for p in self.MD_Projects:
                    for c in self.MD_Projects_Countries[p]:
                        if c in self.Q_COs:
                            self.Q_SalesSegments.append((c,p))
                            self.Q_Projects.append(p)
                            self.Q_RBs.append(self.MD_Projects_RB[p])
            self.Q_Projects = list(set(self.Q_Projects))
            self.Q_SalesSegments = list(set(self.Q_SalesSegments))
            self.Q_SalesSegments.sort()
            self.Q_RBs = list(set(self.Q_RBs))
            self.Q_Countries = list(self.Q_COs)
            t_e = time.time()
            print "Duration: ", "{0:.3f}".format(t_e-t_s), " seconds"

        def commodities():
            print "> Establishing relevant commodities"
            t_s = time.time()
            self.Q_Commodities = [] # set with Commodities-Segment-GMO combinations (specific to Quintiq)
            self.Q_Commodities_SpecCom = {}
            self.Q_Commodities_Segment = {}
            self.Q_Commodities_GMO = {}
            self.Q_Commodities_D_CO = {}
            self.Q_PeggedCommodities = [] # set with Commodity-Segment combinations (specific to Quintiq)
            self.Q_SpecificCommodities = [] # subset of Specific Commodities
            self.Q_CommodityTypes = [] # subset of Commodity Types
            self.Q_FoodGroups = [] # subset of Food Groups
            check = {}
            for c in self.Q_COs:
                self.Q_Commodities_D_CO[c] = []
                check[c] = 0
            for c,p in self.Q_SalesSegments:
                check[c,p] = 0
            self.Q_Demand = {}
            # load tactical demand
            for (c,p,d,sk),(f,nf,t) in self.TacticalDemand.items():
                if c not in self.Q_COs: # Demand for other RB
                    continue
                if c not in self.Demand_Selection.keys():
                    self.weird.append(["Demand Selection not specified yet",c])
                    continue
                elif self.Demand_Selection[c] != (0,1):
                    continue
                if (c,p) not in self.Q_SalesSegments:
                    self.weird.append(["Country-Project combination not recognised",c,p])
                    continue
                if d < self.Q_Start_Planning or d >= add_months(self.Q_Start_Planning,12): # Demand too old or too far away
                    continue
                kp = sk + "_" + c + "_" + p
                self.Q_PeggedCommodities.append(kp)
                self.Q_SpecificCommodities.append(sk)
                self.Q_CommodityTypes.append(self.MD_Commodities_Type[sk])
                self.Q_FoodGroups.append(self.MD_Commodities_Group[sk])
                self.Q_Demand[c,p,d,sk] = (f,nf,t)
                check[c] += 1
                check[c,p] += 1
                for gmo in [" [GMO]"," [Non GMO]"]:
                    if gmo == " [GMO]" and self.GMO[c] == 0: # Country doesn't accept GMO -> no need to create commodity
                        continue
                    k = kp + gmo
                    self.Q_Commodities.append(k)
                    self.Q_Commodities_D_CO[c].append(k)
                    self.Q_Commodities_SpecCom[k] = sk
                    self.Q_Commodities_Segment[k] = (c,p)
                    if gmo == " [GMO]":
                        self.Q_Commodities_GMO[k] = 1
                    else:
                        self.Q_Commodities_GMO[k] = 0
            # load global pipeline
            for (c,p,d,sk),(f,nf,t) in self.PipelineDemand.items():
                if c not in self.Q_COs: # Demand for other RB
                    continue
                if c not in self.Demand_Selection.keys():
                    self.weird.append(["Demand Selection not specified yet",c])
                    continue
                elif self.Demand_Selection[c] != (1,0):
                    continue
                if (c,p) not in self.Q_SalesSegments:
                    self.weird.append(["Country-Project combination not recognised",c,p])
                    continue
                if d < self.Q_Start_Planning or d >= add_months(self.Q_Start_Planning,12): # Demand too old or too far away
                    continue
                kp = sk + "_" + c + "_" + p
                self.Q_PeggedCommodities.append(kp)
                self.Q_SpecificCommodities.append(sk)
                self.Q_CommodityTypes.append(self.MD_Commodities_Type[sk])
                self.Q_FoodGroups.append(self.MD_Commodities_Group[sk])
                self.Q_Demand[c,p,d,sk] = (f,nf,t)
                check[c] += 1
                check[c,p] += 1
                for gmo in [" [GMO]"," [Non GMO]"]:
                    if gmo == " [GMO]" and self.GMO[c] == 0: # Country doesn't accept GMO -> no need to create commodity
                        continue
                    k = kp + gmo
                    self.Q_Commodities.append(k)
                    self.Q_Commodities_D_CO[c].append(k)
                    self.Q_Commodities_SpecCom[k] = sk
                    self.Q_Commodities_Segment[k] = (c,p)
                    if gmo == " [GMO]":
                        self.Q_Commodities_GMO[k] = 1
                    else:
                        self.Q_Commodities_GMO[k] = 0
            # clean up
            self.Q_Commodities = list(set(self.Q_Commodities))
            self.Q_Commodities.sort()
            for c in self.Q_COs:
                self.Q_Commodities_D_CO[c] = list(set(self.Q_Commodities_D_CO[c]))
                self.Q_Commodities_D_CO[c].sort()
            self.Q_PeggedCommodities = list(set(self.Q_PeggedCommodities))
            self.Q_PeggedCommodities.sort()
            self.Q_SpecificCommodities = list(set(self.Q_SpecificCommodities))
            self.Q_SpecificCommodities.sort()
            self.Q_CommodityTypes = list(set(self.Q_CommodityTypes))
            self.Q_CommodityTypes.sort()
            self.Q_FoodGroups = list(set(self.Q_FoodGroups))
            self.Q_FoodGroups.sort()
            # sense check
            for c in sorted(self.Q_COs):
                if check[c] == 0:
                    self.Q_COs.remove(c)
                    self.weird.append(["No demand detected for CO",c])
            for c,p in sorted(self.Q_SalesSegments):
                if c not in self.Q_COs:
                    self.Q_SalesSegments.remove((c,p))
                elif check[c,p] == 0:
                    self.Q_SalesSegments.remove((c,p))
                    self.weird.append(["No demand detected for Country-Project",c,p])
            t_e = time.time()
            print "Duration: ", "{0:.3f}".format(t_e-t_s), " seconds"

        def discharge_ports():
            print "> Establishing relevant discharge ports"
            t_s = time.time()
            self.Q_DPs = []
            self.Q_Commodities_D_DP = {}
            for c in self.Q_COs:
                for dp in self.DischargePorts[c]:
                    self.Q_DPs.append(dp)
                    if dp not in self.Q_Commodities_D_DP.keys():
                        self.Q_Commodities_D_DP[dp] = []
                    for k in self.Q_Commodities_D_CO[c]:
                        self.Q_Commodities_D_DP[dp].append(k)
                    self.Q_Countries.append(self.MD_Locations_Country[dp])
            self.Q_DPs = list(set(self.Q_DPs))
            self.Q_DPs.sort()
            for dp in self.Q_DPs:
                self.Q_Commodities_D_DP[dp] = list(set(self.Q_Commodities_D_DP[dp]))
            t_e = time.time()
            print "Duration: ", "{0:.3f}".format(t_e-t_s), " seconds"

        def loading_ports():
            print "> Establishing relevant load ports"
            t_s = time.time()
            self.Q_LPs = []
            self.Q_Shipping = []
            self.Q_Commodities_D_NDP = {}
            for lane in self.Shipping_Rate.keys(): # lane = (lp,dp)
                lp,dp = lane[0],lane[1]
                if dp in self.Q_DPs:
                    self.Q_LPs.append(lp)
                    self.Q_Shipping.append(lane)
                    if lp not in self.Q_Commodities_D_NDP.keys():
                        self.Q_Commodities_D_NDP[lp] = []
                    for k in self.Q_Commodities_D_DP[dp]:
                        self.Q_Commodities_D_NDP[lp].append(k)
            self.Q_LPs = list(set(self.Q_LPs))
            t_e = time.time()
            print "Duration: ", "{0:.3f}".format(t_e-t_s), " seconds"

        def procurement():
            print "> Establishing relevant procurement options"
            t_s = time.time()
            self.Q_Procurement = []
            self.Q_Procurement_NoDate = []
            self.Q_NDPs = []
            self.Q_SpecificCommodities_S_NDP = {}
            self.Q_SpecificCommodities_GMO = {}
            for proc in self.Procurement_Cost.keys(): # proc = (oc,ndp,com,gmo,date)
                oc,ndp,com,gmo,d,ndp_c = proc[0],proc[1],proc[2],proc[3],proc[4],self.MD_Locations_Country[proc[1]]
                if com not in self.Q_SpecificCommodities: # commodity is irrelevant for this RB
                    continue
                if diff_month(self.Q_Start_Planning,d) > 18: # data is more than one and a half years old
                    continue
                if self.Procurement_Cost[proc] == 0: # price is not known at the moment
                    continue
                if ndp_c in self.Q_COs or ndp in self.Q_LPs:
                    # NB: including local procurement - ndp is already in a destination country
                    self.Q_NDPs.append(ndp)
                    self.Q_Countries.append(oc)
                    self.Q_Countries.append(ndp_c)
                    self.Q_Procurement.append(proc)
                    self.Q_Procurement_NoDate.append((oc,ndp,com,gmo))
                    if ndp not in self.Q_SpecificCommodities_S_NDP.keys():
                        self.Q_SpecificCommodities_S_NDP[ndp] = []
                    if com not in self.Q_SpecificCommodities_GMO.keys():
                        self.Q_SpecificCommodities_GMO[com] = 0
                    self.Q_SpecificCommodities_GMO[com] += gmo  # gmo is a binary index -> if GMO[com] > 1, there is at least 1 GMO option available
                    self.Q_SpecificCommodities_S_NDP[ndp].append(com)
                    if ndp_c in self.Q_COs: # local procurement
                        if ndp not in self.Q_Commodities_D_NDP.keys():
                            self.Q_Commodities_D_NDP[ndp] = []
                        for k in self.Q_Commodities_D_CO[ndp_c]:
                            self.Q_Commodities_D_NDP[ndp].append(k)
            self.Q_NDPs = list(set(self.Q_NDPs))
            self.Q_NDPs.sort()
            self.Q_Countries = list(set(self.Q_Countries))
            self.Q_Countries.sort()
            for ndp in self.Q_NDPs:
                self.Q_SpecificCommodities_S_NDP[ndp] = list(set(self.Q_SpecificCommodities_S_NDP[ndp]))
            self.Q_Procurement_NoDate = list(set(self.Q_Procurement_NoDate))

            # Sense check on commodity list after loading of procurement data
            for k in sorted(self.Q_Commodities):
                sk = self.Q_Commodities_SpecCom[k]
                gmo = self.Q_Commodities_GMO[k]
                if sk not in self.Q_SpecificCommodities_GMO.keys():
                    # no procurement option found for this specific commodity
                    self.weird.append(["Commodity has demand but can't be procured",sk])
                    continue
                if gmo == 1 and self.Q_SpecificCommodities_GMO[sk] == 0:
                    # no procurement option exists for the GMO version of this specific commodity
                    self.Q_Commodities.remove(k)

            # Clean up some of the existing sets
            for co in self.Q_COs:
                self.Q_Commodities_D_CO[co] = [k for k in self.Q_Commodities_D_CO[co] if k in self.Q_Commodities]
            for dp in self.Q_DPs:
                self.Q_Commodities_D_DP[dp] = [k for k in self.Q_Commodities_D_DP[dp] if k in self.Q_Commodities]
            for ndp in self.Q_NDPs:
                self.Q_Commodities_D_NDP[ndp] = [k for k in self.Q_Commodities_D_NDP[ndp] if k in self.Q_Commodities]

            # Filter the procurement options to the most relevant entries (if multiple exist)
            for oc,ndp,com,gmo in self.Q_Procurement_NoDate:
                all_options = [proc for proc in self.Q_Procurement if proc[0] == oc and proc[1] == ndp and proc[2] == com and proc[3] == gmo]
                if len(all_options) > 1: # multiple costs for a (oc,ndp,com,gmo) entry
                    # Find most recent entry
                    dmax = datetime.datetime(1900,1,1)
                    dstart = datetime.datetime(1900,1,1)
                    for option in all_options: # option = (oc,ndp,com,gmo,date)
                        d_u = option[4] # last updated date
                        if d_u > dmax:
                            dmax = d_u
                    for option in all_options: # option = (oc,ndp,com,gmo,date)
                        if option[4] != dstart and option[4] != dmax:
                            self.Q_Procurement.remove(option)
            t_e = time.time()
            print "Duration: ", "{0:.3f}".format(t_e-t_s), " seconds"

        def transport():
            print "> Establishing relevant land transport"
            t_s = time.time()
            self.Q_Transport_DP2CO = [] # lanes for int'l proc
            self.Q_Transport_NDP2CO = [] # lanes for reg'l proc
            for lane in self.Transport_Cost.keys():
                o,d = lane[0],lane[1]
                if o in self.Q_DPs and d in self.Q_COs:
                    self.Q_Transport_DP2CO.append(lane)
                elif o in self.Q_NDPs and d in self.Q_COs:
                    if self.MD_Locations_Country[o] == d:
                        continue # lanes for local proc will be established automatically
                    self.Q_Transport_NDP2CO.append(lane) # tracks overland connections only
                    if o not in self.Q_Commodities_D_NDP.keys():
                        self.Q_Commodities_D_NDP[o] = []
                    for k in self.Q_Commodities_D_CO[d]:
                        self.Q_Commodities_D_NDP[o].append(k)
            for co in self.Q_COs:
                self.Q_Commodities_D_CO[co] = list(set(self.Q_Commodities_D_CO[co]))
                self.Q_Commodities_D_CO[co].sort()
            for ndp in self.Q_NDPs:
                self.Q_Commodities_D_NDP[ndp] = list(set(self.Q_Commodities_D_NDP[ndp]))
                self.Q_Commodities_D_NDP[ndp].sort()
            for dp in self.Q_DPs:
                try: # some DPs are also sources
                    self.Q_Commodities_D_NDP[dp] = list(set(self.Q_Commodities_D_NDP[dp]))
                    self.Q_Commodities_D_NDP[dp].sort()
                except:
                    None
            t_e = time.time()
            print "Duration: ", "{0:.3f}".format(t_e-t_s), " seconds"

        def inventory():
            print "> Establishing relevant inventories"
            t_s = time.time()
            self.Q_Inventory = []
            for key in self.Inventory.keys(): # key = (loc,com,ctry,proj)
                loc, k, c, p = key[0], key[1], key[2], key[3]
                if loc not in (self.Q_COs + self.Q_DPs):
                    continue
                if k not in self.Q_SpecificCommodities:
                    continue
                if p not in self.Q_Projects:
                    continue
                kp = k + "_" + c + "_" + p
                if kp not in self.Q_PeggedCommodities:
                    self.weird.append(["Detected inventory but no demand for",kp])
                    continue
                self.Q_Inventory.append(key)
            t_e = time.time()
            print "Duration: ", "{0:.3f}".format(t_e-t_s), " seconds"

        def gcmf():
            print "> Establishing relevant GCMF options"
            t_s = time.time()
            self.Q_GCMF = []
            self.Q_GCMF_Locations = []
            self.Q_GCMF_Commodities = []
            self.Q_GCMF_Price = {}
            # Stocks
            for zone, loc, com in sorted(self.GCMF.keys()):
                if zone != self.Q_RB and zone in self.MD_Regional_Bureaux:
                    continue # other RB
                if zone != self.Q_RB and loc not in (self.Q_COs + self.Q_DPs):
                    continue # not a relevant location
                if com not in self.Q_SpecificCommodities:
                    self.weird.append(["GCMF inventory but no demand for",com])
                    continue
                if loc not in (self.Q_COs + self.Q_DPs):
                    self.weird.append(["GCMF inventory in unknown location",loc])
                    continue
                if zone != self.Q_RB: # zone = None/null/""
                    if (self.Q_RB,loc,com) in self.GCMF.keys():
                        self.GCMF[self.Q_RB,loc,com][0] += self.GCMF[zone,loc,com][0]
                        self.GCMF[self.Q_RB,loc,com][1] += self.GCMF[zone,loc,com][1]
                    else:
                        self.GCMF[self.Q_RB,loc,com] = self.GCMF[zone,loc,com]
                    zone = self.Q_RB
                self.Q_GCMF.append((loc,com))
                self.Q_GCMF_Locations.append(loc)
                self.Q_GCMF_Commodities.append(com)
            self.Q_GCMF = list(set(self.Q_GCMF))
            self.Q_GCMF.sort()
            self.Q_GCMF_Locations = list(set(self.Q_GCMF_Locations))
            self.Q_GCMF_Locations.sort()
            self.Q_GCMF_Commodities = list(set(self.Q_GCMF_Commodities))
            self.Q_GCMF_Commodities.sort()
            # Price
            for p,sk in self.GCMF_Price.keys():
                usd, d = self.GCMF_Price[p,sk]
                if p not in self.Q_Projects:
                    continue
                if sk not in self.Q_SpecificCommodities:
                    self.weird.append(["GCMF price but no demand for",sk])
                    continue
                if sk not in self.Q_GCMF_Price.keys():
                    self.Q_GCMF_Price[sk] = [usd,d]
                else:
                    if d > self.Q_GCMF_Price[sk][1]: # later entry
                        self.Q_GCMF_Price[sk] = [usd,d]
            t_e = time.time()
            print "Duration: ", "{0:.3f}".format(t_e-t_s), " seconds"

        def forecasts():
            print "> Establishing relevant forecasts"
            t_s = time.time()
            self.Q_Forecast = []
            for key in self.Forecast.keys(): # key = (origin country, ndp, com, gmo, date)
                oc,ndp,com,gmo,d = key[0],key[1],key[2],key[3],key[4]
                if com not in self.Q_SpecificCommodities:
                    continue
                if d <= self.Q_Start_Planning: # old forecast
                    continue
                if ndp not in self.Q_NDPs:
                    if oc in self.Q_COs: # maybe it's a local purchase that we didn't add yet
                        self.weird.append(["Identified forecasts for unknown local procurement option",oc,ndp,com])
                    continue
                self.Q_Forecast.append(key)
            t_e = time.time()
            print "Duration: ", "{0:.3f}".format(t_e-t_s), " seconds"

        def in_kind():
            print "> Establishing relevant In-Kind donations"
            t_s = time.time()
            self.Q_Donation = []
            for don,sk,o,aod,dur,cor,c,p in self.Donation.keys():
                if (c,p) not in self.Q_SalesSegments:
                    continue
                kp = sk + "_" + c + "_" + p
                if kp not in self.Q_PeggedCommodities:
                    self.weird.append(["No demand for donated commodity",kp])
                    continue
                self.Q_Donation.append((don,sk,o,aod,dur,cor,c,p))
                if o not in self.Q_SpecificCommodities_S_NDP.keys():
                    self.Q_SpecificCommodities_S_NDP[o] = [sk]
                    if o not in self.Q_NDPs:
                        self.Q_NDPs.append(o)
                    ndp_c = self.MD_Locations_Country[o]
                    if ndp_c not in self.Q_Countries:
                        self.Q_Countries.append(ndp_c)
                elif sk not in self.Q_SpecificCommodities_S_NDP[o]:
                    self.Q_SpecificCommodities_S_NDP[o].append(sk)
            t_e = time.time()
            print "Duration: ", "{0:.3f}".format(t_e-t_s), " seconds"

        def restrictions():
            print "> Establishing relevant restrictions"
            t_s = time.time()
            self.Q_Sourcing_Restrictions = []
            for oc,k,c,p,d0,d1,ag in self.Sourcing_Restrictions.keys():
                if oc not in self.Q_Countries:
                    continue
                if k not in self.Q_SpecificCommodities and k not in self.Q_CommodityTypes:
                    continue
                if c != None and p != None and (c,p) not in self.Q_SalesSegments:
                    continue
                if d1 != None and d1 < self.Q_Start_Planning:
                    continue
                self.Q_Sourcing_Restrictions.append((oc,k,c,p,d0,d1,ag))
            t_e = time.time()
            print "Duration: ", "{0:.3f}".format(t_e-t_s), " seconds"

        def open_POs():
            print "> Establishing relevant open POs"
            t_s = time.time()
            self.Q_OpenPOs = []
            for rc,p,po,sk,gmo,oc,ndp,d in self.OpenPOs.keys():
                if diff_month(d,self.Q_Start_Planning) < -2:
                    continue
                if p.startswith("S"):
                    if self.GCMF_WBS[p][1] != self.Q_RB:
                        continue # intended for different GCMF zone
                    if sk not in self.Q_GCMF_Commodities:
                        self.Q_GCMF_Commodities.append(sk)
                    if (ndp,sk) not in self.Q_GCMF:
                        self.Q_GCMF.append((ndp,sk))
                else:
                    if (rc,p) not in self.Q_SalesSegments:
                        continue # not a relevant country-project combination
                    gmo_s = " [GMO]" if gmo==1 else " [Non GMO]"
                    k = sk + "_" + rc + "_" + p + gmo_s
                    if k not in self.Q_Commodities:
                        self.weird.append(["Found PO but no demand for",k])
                        continue
                    if (oc,ndp,sk,gmo) not in self.Q_Procurement_NoDate:
                        self.weird.append(["Found PO but no matching procurement option for",oc,ndp,sk,gmo])
                        continue
                self.Q_OpenPOs.append((rc,p,po,sk,gmo,oc,ndp,d))
            t_e = time.time()
            print "Duration: ", "{0:.3f}".format(t_e-t_s), " seconds"

        def unprogrammed():
            print "> Establishing relevant unprogrammed funds"
            t_s = time.time()
            self.Q_Budget = {}
            for rc,p,tm,d,usd in self.Unprogrammed:
                if (rc,p) not in self.Q_SalesSegments:
                    continue
                if tm in ["CD&A","C&V"]:
                    continue # can't use for in-kind
                d = datetime.datetime(d.year,d.month,1)
                if d > self.Q_Start_Planning:
                    continue
                if (rc,p) not in self.Q_Budget.keys():
                    self.Q_Budget[rc,p] = 0
                self.Q_Budget[rc,p] += usd
            t_e = time.time()
            print "Duration: ", "{0:.3f}".format(t_e-t_s), " seconds"

        def bbd():
            print "> Establishing relevant BBDs"
            t_s = time.time()
            self.Q_Supply = {}
            self.Q_BBDinv = {}
            self.Q_BudgetBonus = {}

            for key in self.Q_Inventory:
                sk,c,p = key[1],key[2],key[3]
                inv,itr = self.Inventory[key]
                if (c,p,sk) not in self.Q_Supply.keys():
                    self.Q_Supply[c,p,sk] = 0
                self.Q_Supply[c,p,sk] += inv + itr
            for c,p,po,sk,gmo,oc,ndp,d in self.Q_OpenPOs:
                if p.startswith("S"):
                    continue # GCMF
                q = self.OpenPOs[c,p,po,sk,gmo,oc,ndp,d]
                pod = add_months(d,-2)
                if diff_month(self.Q_Start_Planning,pod) <= 0: # PO starts during planning horizon
                    continue
                if (c,p,sk) not in self.Q_Supply.keys():
                    self.Q_Supply[c,p,sk] = 0
                self.Q_Supply[c,p,sk] += q

            for (c,p,d,sk),(f,nf,t) in self.Q_Demand.items():
                if d <= add_months(self.Q_Start_Planning,6):
                    d = self.Q_Start_Planning
                else:
                    d = add_months(d,-6)
                if (c,p,sk,d) not in self.Q_BBDinv.keys():
                    self.Q_BBDinv[c,p,sk,d] = 0
                self.Q_BBDinv[c,p,sk,d] += f

            for c,p,sk in self.Q_Supply.keys():
                s = self.Q_Supply[c,p,sk]
                try:
                    d = self.Q_BBDinv[c,p,sk,self.Q_Start_Planning]
                except:
                    self.weird.append(["Initial supply but no demand for",c,p,sk,s])
                    d = 0
                if s > d:
                    td = sum(dem[1] for dem in self.Q_BBDinv.items() if dem[0][0]==c and dem[0][1]==p and dem[0][2]==sk)
                    if s > td:
                        self.weird.append(["Existing inventory > 12 months of demand for",c,p,sk,s,td])
                        if (c,p) not in self.Q_BudgetBonus.keys():
                            self.Q_BudgetBonus[c,p] = 0
                        self.Q_BudgetBonus[c,p] += (s-td) * self.StaticFCR[c,p,sk]
                    else:
                        self.weird.append(["Existing inventory > 6 months of demand for",c,p,sk,s,d])
                    self.Q_BBDinv[c,p,sk,self.Q_Start_Planning] = 0
                else:
                    self.Q_BBDinv[c,p,sk,self.Q_Start_Planning] = d - s
            t_e = time.time()
            print "Duration: ", "{0:.3f}".format(t_e-t_s), " seconds"

        # Execute
        self.weird = []
        projects()
        commodities()
        discharge_ports()
        loading_ports()
        procurement()
        transport()
        inventory()
        gcmf()
        forecasts()
        in_kind()
        restrictions()
        open_POs()
        unprogrammed()
        bbd()
        print "> Exporting weird occurences to 'Debugging.xlsx'"
        self.print_to_file(self.weird,"Weird occurences (mapping)",["Type","Details"])

    def export_Quintiq(self):
        '''
        Export data to Quintiq MP format
        '''

        def Periods():
            print "> MP_Periods"
            t_s = time.time()
            path0 = os.path.join(self.data_dir,"MP_Periods.xlsx")
            path1 = os.path.join(self.dest_dir,"MP_Periods.xlsx")
            wb = xl.load_workbook(filename = path0, read_only=False)

            ws = wb.get_sheet_by_name("PeriodParameters_MP")
            # row = [ID, StartOfPlanning]
            ws.append([1,self.Q_Start_Planning])


            ws = wb.get_sheet_by_name("PeriodSpecifications_MP")
            # row = FrozenFuturePeriod	ID	NrOfFuturePeriod	NrOfHistoricalPeriod	NrOfTimeUnit	PeriodAlignment	TimeUnit
            ws.append([0,"Months",14,0,1,self.Q_Start_Horizon,"Month"])

            ws = wb.get_sheet_by_name("PeriodBudget_MP")
            # row = [Start, End, ExpectedBudget]
            budget = {}
            dmax = datetime.datetime(2000,1,1)
            for (c,p,d,sk),(f,nf,t) in self.Q_Demand.items():
                if d > dmax:
                    dmax = d
                if d not in budget.keys():
                    budget[d] = 0
                budget[d] += f * self.StaticFCR[c,p,sk]
            for t in range(diff_month(dmax,self.Q_Start_Planning)+1):
                d = add_months(self.Q_Start_Planning,t)
                ws.append([d,add_months(d,1),budget[d]])
            # period budget to be redefined through separate Excel file

            wb.save(filename = path1)
            t_e = time.time()
            print "Duration: ", "{0:.3f}".format(t_e-t_s), " seconds"

        def UnitOfMeasures():
            print "> MP_UnitOfMeasures"
            t_s = time.time()
            path0 = os.path.join(self.data_dir,"MP_UnitOfMeasures.xlsx")
            path1 = os.path.join(self.dest_dir,"MP_UnitOfMeasures.xlsx")
            wb = xl.load_workbook(filename = path0, read_only=False)

            ws = wb.get_sheet_by_name("BaseConversionFactors_MP")
            # row = Factor	IsEnabled	ProductID	SourceUnitOfMeasureName	TargetUnitOfMeasureName
            for k in self.Q_Commodities:
                sk = self.Q_Commodities_SpecCom[k]
                if sk not in self.CommodityIntake.keys():
                    self.weird.append(("No Container conversion rate found",sk))
                    continue
                ws.append([self.CommodityIntake[sk],"TRUE",k,"Container","MT"])

            wb.save(filename = path1)
            t_e = time.time()
            print "Duration: ", "{0:.3f}".format(t_e-t_s), " seconds"

        def Units():
            print "> MP_Units"
            t_s = time.time()
            path0 = os.path.join(self.data_dir,"MP_Units.xlsx")
            path1 = os.path.join(self.dest_dir,"MP_Units.xlsx")
            wb = xl.load_workbook(filename = path0, read_only=False)
            ws = wb.active
            # row = [GroupName	HasCapacitySmoothing	HasMaximumQuantity	HasUserQuantityToProcess	ID	IsBulk	IsInKind
            # IsIncludedInSupplySpecification	IsOverlapCapacitySmoothing	IsPlannedInfinite	IsSmoothCapacityOverAllPeriods
            # LotSize	MaximumQuantity	MinimumQuantity	Name	Notes	ParentUnitID	StartDate	UnitOfMeasureName	UseGroupIcon
            # UserCapacitySmoothingLength	UserIconName	UserQuantityToProcess	IsSupplierUnit   IsGCMF]

            r = 13
            for c in sorted(self.Q_Countries):
                r += 1
                ws.append(["3650 days",0,"Quantity","USD","",r,datetime.datetime(9999, 12, 31),"","","", \
                           "Country","FALSE","FALSE","FALSE",c,"FALSE","FALSE","TRUE","FALSE","FALSE","FALSE", \
                           0,0,0,c,"","COUNTRIES",datetime.datetime(1900, 1, 1),"MT","TRUE",2,"","out","TRUE","FALSE"])
            r = 200
            for dp in sorted(self.Q_DPs):
                # port unit (aggregation)
                r += 1
                ws.append(["3650 days",0,"Infinite","USD","",r,datetime.datetime(9999, 12, 31),"","","", \
                           "Discharge Port (In)","FALSE","FALSE","FALSE",dp,"FALSE","FALSE","TRUE","FALSE","FALSE","FALSE", \
                           0,0,0,dp,"","PORTS",datetime.datetime(1900, 1, 1),"MT","FALSE",2,"CUBES","out","FALSE","FALSE"])
                # port unit (sub-units)
                icon = {"Bulk": "CUBES_BLUE", "Break-Bulk": "CUBES_YELLOW", "Container": "CUBES_GREEN"}
                # bulk = {"Bulk": "TRUE", "Break-Bulk": "TRUE", "Container": "FALSE"}
                for t in self.MD_ShippingTypes:
                    r += 1
                    name = dp + " [" + t + "]"
                    ws.append(["3650 days",0,"Quantity","USD","",r,datetime.datetime(9999, 12, 31),"","","", \
                               "Discharge Port (In)","FALSE","FALSE","FALSE",name,"FALSE","FALSE","TRUE","FALSE","FALSE","FALSE", \
                               0,0,0,name,"",dp,datetime.datetime(1900, 1, 1),"MT","FALSE",2,icon[t],"out","FALSE","FALSE"])

            wb.save(filename = path1)
            t_e = time.time()
            print "Duration: ", "{0:.3f}".format(t_e-t_s), " seconds"

        def StockingPoints():
            print "> MP_StockingPoints"
            t_s = time.time()
            path0 = os.path.join(self.data_dir,"MP_StockingPoints.xlsx")
            path1 = os.path.join(self.dest_dir,"MP_StockingPoints.xlsx")
            wb = xl.load_workbook(filename = path0, read_only=False)
            ws = wb.active
            # row = CurrencyID	DisplayIndex	End	GISCity	GISCountryCode	GISPostalCode
            # GroupName	ID	IsPlannedInfinite	Name	Notes	Start	UnitID
            # UnitOfMeasureName	UseGroupIcon	UserIconName

            r = 0
            for co in sorted(self.Q_COs):
                name = co + " [D]"
                r += 1
                ws.append(["USD",r,datetime.datetime(9999, 12, 31),"","","","Country Office",name, \
                           "FALSE",name,"",datetime.datetime(1900, 1, 1),co,"MT","TRUE",""])
                name = co + " [WH]"
                r += 1
                ws.append(["USD",r,datetime.datetime(9999, 12, 31),"","","","Country Office",name, \
                           "FALSE",name,"",datetime.datetime(1900, 1, 1),co,"MT","FALSE","HOUSE"])
            for dp in sorted(self.Q_DPs):
                r += 1
                ws.append(["USD",r,datetime.datetime(9999, 12, 31),"","","","Discharge Port (Out)",dp, \
                           "FALSE",dp,"",datetime.datetime(1900, 1, 1),self.MD_Locations_Country[dp],"MT","TRUE",""])
                icon = {"Bulk": "CUBES_BLUE", "Break-Bulk": "CUBES_YELLOW", "Container": "CUBES_GREEN"}
                for st in ["Bulk","Break-Bulk","Container"]:
                    r += 1
                    name = dp + " [" + st + "]"
                    ws.append(["USD",r,datetime.datetime(9999, 12, 31),"","","","Discharge Port (In)",name, \
                               "FALSE",name,"",datetime.datetime(1900, 1, 1),self.MD_Locations_Country[dp],"MT","FALSE",icon[st]])
            for ndp in sorted(self.Q_NDPs):
                if ndp in self.Q_DPs:
                    continue # stocking point already created
                r += 1
                c = self.MD_Locations_Country[ndp]
                ws.append(["USD",r,datetime.datetime(9999, 12, 31),"","","","Named Delivery Place",ndp, \
                           "FALSE",ndp,"",datetime.datetime(1900, 1, 1),c,"MT","TRUE",""])
            if len(self.Q_Donation) > 0:
                r += 1
                ws.append(["USD",r,datetime.datetime(9999, 12, 31),"","","","USAID","In-Kind Donation", \
                           "FALSE","In-Kind Donation","",datetime.datetime(1900, 1, 1),"","MT","TRUE",""])
            r += 1
            ws.append(["USD",r,datetime.datetime(9999, 12, 31),"","","","","Virtual Warehouse [BBD]", \
                        "TRUE","Virtual Warehouse [BBD]","",datetime.datetime(1900, 1, 1),"","MT","FALSE","CABINET"])
            for c in self.Q_Countries:
                wh = "Virtual Warehouse [" + c + "]"
                r += 1
                ws.append(["USD",r,datetime.datetime(9999, 12, 31),"","","","",wh, \
                           "TRUE",wh,"",datetime.datetime(1900, 1, 1),"","MT","FALSE","CABINET"])

            wb.save(filename = path1)
            t_e = time.time()
            print "Duration: ", "{0:.3f}".format(t_e-t_s), " seconds"

        def Products():
            print "> MP_Products"
            t_s = time.time()
            path0 = os.path.join(self.data_dir,"MP_Products.xlsx")
            path1 = os.path.join(self.dest_dir,"MP_Products.xlsx")
            wb = xl.load_workbook(filename = path0, read_only=False)
            ws = wb.get_sheet_by_name("Products_MP")
            # row = DisplayIndex	End	HasShelfLife	ID	IconName	IsByProduct	Name	Notes
            # ParentID	RecipeName	ShelfLife	Start	UnitOfMeasureName	IsGMO	PriorityName	CanShipAsBulk

            r = 0
            # Level 0 - Food Groups
            for fg in sorted(self.Q_FoodGroups):
                bulk = "FALSE"
                for kt in self.Q_CommodityTypes:
                    if self.MD_CommodityTypes_Group[kt] == fg:
                        if kt in self.Bulk_Commodity["Bulk"]:
                            bulk = "TRUE"
                            break
                r += 1
                ws.append([r,datetime.datetime(9999, 12, 31),"FALSE",fg,"","FALSE",fg,"","","",0,datetime.datetime(1900, 1, 1),"MT","FALSE","None",bulk])
            # Level 1 - Generic Commodity Names
            for kt in sorted(self.Q_CommodityTypes):
                bulk = "TRUE" if kt in self.Bulk_Commodity["Bulk"] else "FALSE"
                r += 1
                ws.append([r,datetime.datetime(9999, 12, 31),"FALSE",kt,"","FALSE",kt,"",self.MD_CommodityTypes_Group[kt],"",0,datetime.datetime(1900, 1, 1),"MT","FALSE","None",bulk])
            # Level 2 - Specific Commodity Names
            for sk in sorted(self.Q_SpecificCommodities):
                kt = self.MD_Commodities_Type[sk]
                bulk = "TRUE" if kt in self.Bulk_Commodity["Bulk"] else "FALSE"
                r += 1
                ws.append([r,datetime.datetime(9999, 12, 31),"FALSE",sk,"","FALSE",sk,"",self.MD_Commodities_Type[sk],"",0,datetime.datetime(1900, 1, 1),"MT","FALSE","None",bulk])
            # Level 3 - Pegged Commodities
            for kp in sorted(self.Q_PeggedCommodities): # kp = WHEAT_MALI_200719
                sk = kp[:kp.find("_")]
                cp = kp[kp.find("_")+1:]
                c = cp[:cp.find("_")]
                p = cp[cp.find("_")+1:]
                kt = self.MD_Commodities_Type[sk]
                bulk = "TRUE" if kt in self.Bulk_Commodity["Bulk"] else "FALSE"
                if (sk,p) in self.Commodity_Priority.keys():
                    ratio = self.Commodity_Priority[sk,p]
                    ratio = max(1,math.ceil(ratio*10))
                    priority = "Q" + str(int(ratio))
                else:
                    priority = "Q1"
                r += 1
                ws.append([r,datetime.datetime(9999, 12, 31),"FALSE",kp,"","FALSE",kp,"",sk,"",0,datetime.datetime(1900, 1, 1),"MT","FALSE",priority,bulk])
            # Level 4 - GMO & Pegged Commodities
            for k in sorted(self.Q_Commodities):
                sk = self.Q_Commodities_SpecCom[k]
                c,p = self.Q_Commodities_Segment[k]
                kp = sk + "_" + c + "_" + p
                if self.Q_Commodities_GMO[k] == 1:
                    gmo = "TRUE"
                    gmo_s = " [GMO]"
                else:
                    gmo = "FALSE"
                    gmo_s = " [Non GMO]"
                kt = self.MD_Commodities_Type[sk]
                bulk = "TRUE" if kt in self.Bulk_Commodity["Bulk"] else "FALSE"
                if (sk,p) in self.Commodity_Priority.keys():
                    ratio = self.Commodity_Priority[sk,p]
                    ratio = max(1,math.ceil(ratio*10))
                    priority = "Q" + str(int(ratio))
                else:
                    priority = "Q1"
                    self.weird.append(("No Commodity priority",sk,p))
                r += 1
                ws.append([r,datetime.datetime(9999, 12, 31),"FALSE",k,"","FALSE",k,"",kp,"",160,datetime.datetime(1900, 1, 1),"MT",gmo,priority,bulk])
            # Level 3 & 4 - GCMF Pegged Commodities
            for k in sorted(self.Q_GCMF_Commodities):
                kt = self.MD_Commodities_Type[k]
                bulk = "TRUE" if kt in self.Bulk_Commodity["Bulk"] else "FALSE"
                com3 = k + "_GCMF"
                r += 1
                ws.append([r,datetime.datetime(9999, 12, 31),"FALSE",com3,"","FALSE",com3,"",k,"",0,datetime.datetime(1900, 1, 1),"MT","FALSE","None",bulk])
                com4 = com3 + " [Non GMO]"
                r += 1
                ws.append([r,datetime.datetime(9999, 12, 31),"FALSE",com4,"","FALSE",com4,"",com3,"",160,datetime.datetime(1900, 1, 1),"MT","FALSE","None",bulk])

            wb.save(filename = path1)
            t_e = time.time()
            print "Duration: ", "{0:.3f}".format(t_e-t_s), " seconds"

        def Routings():
            print "> MP_Routings"
            t_s = time.time()
            path0 = os.path.join(self.data_dir,"MP_Routings.xlsx")

            self.Q_Operations = {}
            for baseline in [0,1]:
                if baseline == 0:
                    path1 = os.path.join(self.dest_dir,"MP_Routings.xlsx")
                else:
                    path1 = os.path.join(self.dest_dir,"Baseline\MP_Routings.xlsx")
                wb = xl.load_workbook(filename = path0, read_only=False)

                ws0 = wb.get_sheet_by_name("Routings_MP")
                # row = End	ID	IsEnabled	Name	Start
                ws1 = wb.get_sheet_by_name("RoutingSteps_MP")
                # row = Name	RoutingID	SequenceNumber
                ws2 = wb.get_sheet_by_name("Operations_MP")
                # row = HasUserLotSize	HasUserMaximumQuantity	ID	IsEnabled	LotSize	MaximumQuantity
                # MinimumQuantity	RoutingID	RoutingStepName	StandardDeviationLeadTimeOperation	Throughput	UnitID	UserLeadTime
                ws3 = wb.get_sheet_by_name("OperationBOMs_MP")
                # row = InputGroupID	IsExcluded	IsInput	MaxQuantityInGroup	MinQuantityInGroup
                # OperationID	OperationInputSetName	ProductID	Quantity	StockingPointID
                ws4 = wb.get_sheet_by_name("RoutingStepInPeriod")
                # row = ActualFCR	DynamicFCR	StaticFCR	PeriodEnd	PeriodStart	RoutingStepName	RoutingID

                # Procurement options
                for key in self.Q_Procurement_NoDate: # key = (oc,ndp,com,gmo)
                    oc,ndp,sk,gmo = key[0],key[1],key[2],key[3]
                    ic = self.MD_Locations_Country[ndp] # incoterm country
                    if gmo == 1:
                        gmo_s = " [GMO]"
                    else:
                        gmo_s = " [Non GMO]"
                    coms = [k for k in self.Q_Commodities_D_NDP[ndp] if self.Q_Commodities_SpecCom[k] == sk and self.Q_Commodities_GMO[k] == gmo]
                    if ndp in self.Q_DPs:
                        for k in self.Q_Commodities_D_DP[ndp]:
                            if self.Q_Commodities_SpecCom[k] == sk and self.Q_Commodities_GMO[k] == gmo:
                                coms.append(k)
                    coms = list(set(coms))
                    if len(coms) == 0:
                        continue # no projects have access and demand to this purchasing option
                    rname = "Buy " + sk + gmo_s + " from " + oc + " at " + ndp
                    ws0.append([datetime.datetime(9999, 12, 31),rname,"TRUE",rname,datetime.datetime(1900, 1, 1)])
                    i = 0
                    for k in coms:
                        c,p = self.Q_Commodities_Segment[k]
                        oname = "Buy " + k + " from " + oc + " at " + ndp
                        rstep = oc + "_" + c + "_" + p
                        i += 1
                        ws1.append([rstep,rname,i])
                        ws2.append(["FALSE","FALSE",oname,"TRUE",0,0,0,rname,rstep,"0:00:00.00",1,oc,"64 days"])
                        ws3.append([1,"FALSE","FALSE",1,1,oname,"",k,1,ndp]) # output commodity
                        vk = sk + "_" + c + "_" + p + " [Non GMO]"
                        ws3.append([1,"FALSE","TRUE",1,1,oname,"",vk,1,"Virtual Warehouse [BBD]"]) # input (virtual) commodity
                        if baseline == 1:
                            wh = "Virtual Warehouse [" + ic + "]"
                            ws3.append([1,"FALSE","TRUE",1,1,oname,"",vk,1,wh]) # input (virtual) commodity
                        for t in range(12+diff_month(self.Q_Start_Planning,self.Q_Start_Horizon)):
                            d0 = add_months(self.Q_Start_Horizon,t)
                            d1 = add_months(self.Q_Start_Horizon,t+1)
                            ws4.append(["",self.StaticFCR[c,p,sk],self.StaticFCR[c,p,sk],d1,d0,rstep,rname])
                        self.Q_Operations[oname] = (k,vk,ndp,"Procurement")
                # GCMF procurement
                for loc,sk in self.Q_GCMF:
                    rname = "Buy " + sk + " from GCMF at " + loc
                    ws0.append([datetime.datetime(9999, 12, 31),rname,"TRUE",rname,datetime.datetime(1900, 1, 1)])
                    coms = [k for k in self.Q_Commodities if self.Q_Commodities_SpecCom[k] == sk and self.Q_Commodities_GMO[k] == 0]
                    gk = sk + "_GCMF [Non GMO]"
                    for k in coms:
                        c,p = self.Q_Commodities_Segment[k]
                        oname = "Buy " + k + " from GCMF at " + loc
                        if loc in self.Q_COs:
                            if loc != c:
                                continue
                            wh = loc + " [WH]"
                        else:
                            wh = loc
                        rstep = "GCMF_" + c + "_" + p
                        i += 1
                        ws1.append([rstep,rname,i])
                        ws2.append(["FALSE","FALSE",oname,"TRUE",0,0,0,rname,rstep,"0:00:00.00",1,"GCMF","15 days"])
                        ws3.append([1,"FALSE","FALSE",1,1,oname,"",k,1,wh])
                        ws3.append([1,"FALSE","TRUE",1,1,oname,"",gk,1,wh])
                        vk = sk + "_" + c + "_" + p + " [Non GMO]"
                        ws3.append([1,"FALSE","TRUE",1,1,oname,"",vk,1,"Virtual Warehouse [BBD]"]) # input (virtual) commodity
                        for t in range(12+diff_month(self.Q_Start_Planning,self.Q_Start_Horizon)):
                            d0 = add_months(self.Q_Start_Horizon,t)
                            d1 = add_months(self.Q_Start_Horizon,t+1)
                            ws4.append(["",self.StaticFCR[c,p,sk],self.StaticFCR[c,p,sk],d1,d0,rstep,rname])
                        self.Q_Operations[oname] = (k,gk,wh,"GCMF")
                # In-Kind 'procurement'
                rnumber = {}
                ops = []
                for don,sk,o,aod,dur,cor,c,p in self.Q_Donation:
                    rname = "Donation of " + sk + " by " + don + " at " + o
                    ic = self.MD_Locations_Country[o]
                    if (sk,don,o) not in rnumber.keys():
                        ws0.append([datetime.datetime(9999, 12, 31),rname,"TRUE",rname,datetime.datetime(1900, 1, 1)])
                        rnumber[sk,don,o] = 1
                    k = sk + "_" + c + "_" + p + " [Non GMO]"
                    oname = "Donation of " + k + " by " + don + " at " + o
                    if oname in ops:
                        continue
                    else:
                        ops.append(oname)
                    rstep = don + "_" + c + "_" + p
                    i += 1
                    ws1.append([rstep,rname,i])
                    ws2.append(["FALSE","FALSE",oname,"TRUE",0,0,0,rname,rstep,"0:00:00.00",1,"DONATION","1 days"])
                    if baseline == 1:
                        wh = "Virtual Warehouse [" + ic + "]"
                        ws3.append([1,"FALSE","TRUE",1,1,oname,"",k,1,wh]) # input (virtual) commodity
                    else:
                        ws3.append([1,"FALSE","TRUE",1,1,oname,"",k,1,"In-Kind Donation"])
                    ws3.append([1,"FALSE","TRUE",1,1,oname,"",k,1,"Virtual Warehouse [BBD]"]) # input (virtual) commodity
                    ws3.append([1,"FALSE","FALSE",1,1,oname,"",k,1,o])
                    for t in range(12+diff_month(self.Q_Start_Planning,self.Q_Start_Horizon)):
                        d0 = add_months(self.Q_Start_Horizon,t)
                        d1 = add_months(self.Q_Start_Horizon,t+1)
                        ws4.append(["",self.StaticFCR[c,p,sk],self.StaticFCR[c,p,sk],d1,d0,rstep,rname])
                    self.Q_Operations[oname] = (k,k,o,"Donation")
                # Unloading operations
                rnumber = {}
                gcmf_check =  []
                for dp in self.Q_DPs:
                    if dp in self.Handling_Dur.keys():
                        dur = self.Handling_Dur[dp]
                    else:
                        dur = 0
                        self.weird.append(("No handling time for port",dp))
                    for k in self.Q_Commodities_D_DP[dp]:
                        sk,gmo = self.Q_Commodities_SpecCom[k], self.Q_Commodities_GMO[k]
                        c,p = self.Q_Commodities_Segment[k]
                        if gmo == 1:
                            gmo_s = " [GMO]"
                        else:
                            gmo_s = " [Non GMO]"
                        for t in self.MD_ShippingTypes:
                            dpt = dp + " [" + t + "]"
                            rname = "Unloading " + sk + gmo_s + " in " + dpt
                            if (sk,dpt,gmo) not in rnumber.keys():
                                rnumber[sk,dpt,gmo] = 0
                                ws0.append([datetime.datetime(9999, 12, 31),rname,"TRUE",rname,datetime.datetime(1900, 1, 1)])
                            rstep = c + "_" + p + "_" + dpt
                            rnumber[sk,dpt,gmo] += 1
                            ws1.append([rstep,rname,rnumber[sk,dpt,gmo]])
                            oname = "Unloading " + k + " in " + dpt
                            ws2.append(["FALSE","FALSE",oname,"TRUE",0,0,0,rname,rstep,"0:00:00.00",1,dpt,str(dur) + " days"])
                            ws3.append([1,"FALSE","TRUE",1,1,oname,"",k,1,dpt])
                            ws3.append([1,"FALSE","FALSE",1,1,oname,"",k,1,dp])
                            if dp in self.GCMF_Ports and self.GCMF_Commodity[c,p,sk]==1 and gmo==0 and (dpt,sk) not in gcmf_check:
                                rstep = "GCMF_" + dpt
                                ws1.append([rstep,rname,1])
                                gk = sk + "_GCMF [Non GMO]"
                                oname = "Unloading " + gk + " in " + dpt
                                ws2.append(["FALSE","FALSE",oname,"TRUE",0,0,0,rname,rstep,"0:00:00.00",1,dpt,str(dur) + " days"])
                                ws3.append([1,"FALSE","TRUE",1,1,oname,"",gk,1,dpt])
                                ws3.append([1,"FALSE","FALSE",1,1,oname,"",gk,1,dp])
                                gcmf_check.append((dpt,sk))
                wb.save(filename = path1)

            t_e = time.time()
            print "Duration: ", "{0:.3f}".format(t_e-t_s), " seconds"

        def Lanes():
            print "> MP_Lanes"
            t_s = time.time()
            path0 = os.path.join(self.data_dir,"MP_Lanes.xlsx")
            path1 = os.path.join(self.dest_dir,"MP_Lanes.xlsx")
            wb = xl.load_workbook(filename = path0, read_only=False)

            ws0 = wb.get_sheet_by_name("Lanes_MP")
            # row = End	ID	IsEnabled	Name	StandardDeviationLeadTimeLane	Start	UnitID	UserLeadTime
            ws1 = wb.get_sheet_by_name("ProductInLanes_MP")
            # row = IsExcluded	LaneID	ProductID
            ws2 = wb.get_sheet_by_name("LaneLegs_MP")
            # row = DestinationStockingPointID	End	HasStandardDeviationLeadTimeLaneLeg	HasUserLeadTime	IsEnabled
            # LaneID	Name	OriginStockingPointID	StandardDeviationLeadTimeLaneLeg	Start	UserLeadTime
            r = 0
            # NDP - DP lanes (shipping)
            self.Q_Bulk_Coms = {}
            check_gcmf = []
            for lane in sorted(self.Q_Shipping):
                lp,dp = lane[0],lane[1]
                if lp not in self.Q_DPs:
                    if lp not in self.Q_SpecificCommodities_S_NDP.keys(): # LP not a source for anything we need
                        self.Q_Shipping.remove(lane)
                        continue
                    valid = 0
                    for k in self.Q_Commodities_D_DP[dp]:
                        sk = self.Q_Commodities_SpecCom[k]
                        if sk in self.Q_SpecificCommodities_S_NDP[lp]:
                            valid = 1
                            break
                    if valid == 0: # LP can't ship any commodities that this DP needs
                        self.Q_Shipping.remove(lane)
                        continue
                if lane in self.Shipping_Duration.keys():
                    lt = self.Shipping_Duration[lane]*24
                else:
                    lt = 40*24 # if the lead time is 0 I can't use the lane :(
                    self.weird.append(("No lead time found for lane",lp,dp))
                for t in self.MD_ShippingTypes:
                    name = "Ship from " + lp + " to " + dp + " in " + t
                    dpt = dp + " [" + t + "]"
                    if t == "Container":
                        activated = "TRUE"
                    elif ("",dp,t) in self.Bulk_Lanes:
                        activated = "TRUE"
                    elif (lp,dp,t) in self.Bulk_Lanes:
                        activated = "TRUE"
                    else:
                        activated = "FALSE"
                    if t == "Bulk":
                        routing = {}
                        for k in self.Q_Commodities_D_DP[dp]:
                            sk = self.Q_Commodities_SpecCom[k]
                            c,p = self.Q_Commodities_Segment[k]
                            gmo = self.Q_Commodities_GMO[k]
                            gmo_s = " [GMO]" if gmo == 1 else " [Non GMO]"
                            kt = self.MD_Commodities_Type[sk]
                            if kt not in self.Bulk_Commodity["Bulk"]:
                                continue
                            if lp not in self.GCMF_Ports and dp not in self.GCMF_Ports and self.GCMF_Commodity[c,p,sk]==1 and self.GCMF_Priority=="Hard":
                                continue
                            if lp in self.Q_DPs:
                                name_b = name + " (" + sk + gmo_s + ")"
                                if (sk,gmo) not in routing.keys():
                                    ws0.append([datetime.datetime(9999, 12, 31),name_b,activated,name_b,"00:00:00",datetime.datetime(1900, 1, 1),t,str(lt)+":00:00.00"])
                                    r += 1
                                    ws2.append([dpt,datetime.datetime(9999, 12, 31),"FALSE","FALSE","TRUE",name_b,r,lp,"0:00:00.00",datetime.datetime(1900, 1, 1),str(lt)+":00:00.00"])
                                    routing[sk,gmo] = 1
                                    if (lp,dp) not in self.Q_Bulk_Coms.keys():
                                        self.Q_Bulk_Coms[lp,dp] = []
                                    self.Q_Bulk_Coms[lp,dp].append((sk,gmo))
                                ws1.append(["FALSE",name_b,k])
                                if (lp,dp,t,sk,gmo) not in check_gcmf:
                                    check_gcmf.append((lp,dp,t,sk,gmo))
                                    if dp in self.GCMF_Ports and sk in self.Q_GCMF_Commodities:
                                        gk = sk + "_GCMF" + gmo_s
                                        ws1.append(["FALSE",name_b,gk])
                            else:
                                if sk in self.Q_SpecificCommodities_S_NDP[lp]:
                                    name_b = name + " (" + sk + gmo_s + ")"
                                    if (sk,gmo) not in routing.keys():
                                        ws0.append([datetime.datetime(9999, 12, 31),name_b,activated,name_b,"00:00:00",datetime.datetime(1900, 1, 1),t,str(lt)+":00:00.00"])
                                        r += 1
                                        ws2.append([dpt,datetime.datetime(9999, 12, 31),"FALSE","FALSE","TRUE",name_b,r,lp,"0:00:00.00",datetime.datetime(1900, 1, 1),str(lt)+":00:00.00"])
                                        routing[sk,gmo] = 1
                                        if (lp,dp) not in self.Q_Bulk_Coms.keys():
                                            self.Q_Bulk_Coms[lp,dp] = []
                                        self.Q_Bulk_Coms[lp,dp].append((sk,gmo))
                                    ws1.append(["FALSE",name_b,k])
                                    if (lp,dp,t,sk,gmo) not in check_gcmf:
                                        check_gcmf.append((lp,dp,t,sk,gmo))
                                        if dp in self.GCMF_Ports and sk in self.Q_GCMF_Commodities:
                                            gk = sk + "_GCMF" + gmo_s
                                            ws1.append(["FALSE",name_b,gk])
                    elif t == "Break-Bulk" or t == "Container":
                        ws0.append([datetime.datetime(9999, 12, 31),name,activated,name,"00:00:00",datetime.datetime(1900, 1, 1),t,str(lt)+":00:00.00"])
                        for k in self.Q_Commodities_D_DP[dp]:
                            sk = self.Q_Commodities_SpecCom[k]
                            kt = self.MD_Commodities_Type[sk]
                            c,p = self.Q_Commodities_Segment[k]
                            if t == "Break-Bulk" and kt not in self.Bulk_Commodity["Break-Bulk"]:
                                continue
                            if lp not in self.GCMF_Ports and dp not in self.GCMF_Ports and self.GCMF_Commodity[c,p,sk]==1 and self.GCMF_Priority=="Hard":
                                continue
                            if lp in self.Q_DPs:
                                ws1.append(["FALSE",name,k])
                                if (lp,dp,t,sk) not in check_gcmf:
                                    check_gcmf.append((lp,dp,t,sk))
                                    if dp in self.GCMF_Ports and sk in self.Q_GCMF_Commodities:
                                        gk = sk + "_GCMF" + gmo_s
                                        ws1.append(["FALSE",name,gk])
                            else:
                                if sk in self.Q_SpecificCommodities_S_NDP[lp]:
                                    ws1.append(["FALSE",name,k])
                                    if (lp,dp,t,sk) not in check_gcmf:
                                        check_gcmf.append((lp,dp,t,sk))
                                        if dp in self.GCMF_Ports and sk in self.Q_GCMF_Commodities:
                                            gk = sk + "_GCMF" + gmo_s
                                            ws1.append(["FALSE",name,gk])
                        r += 1
                        ws2.append([dpt,datetime.datetime(9999, 12, 31),"FALSE","FALSE","TRUE",name,r,lp,"0:00:00.00",datetime.datetime(1900, 1, 1),str(lt)+":00:00.00"])
                    else:
                        self.weird.append(("Shipping type not accounted for",t))

            # DP - CO lanes (inland/overland transport)
            for dp,c in self.Q_Transport_DP2CO:
                if (dp,c) in self.Transport_Dur.keys():
                    lt = self.Transport_Dur[dp,c]*24
                else:
                    lt = 40*24 # if the lead time is 0 I can't use the lane :(
                    self.weird.append(("No lead time found for lane",dp,c))
                if self.MD_Locations_Country[dp] == self.MD_Locations_Country[c]:
                    unit = "Road - Inland"
                else:
                    unit = "Road - Overland"
                projects = [p for (co,p) in self.Q_SalesSegments if co == c] # Identify relevant projects
                for p in projects:
                    name = "Truck from " + dp + " to " + c + " for " + p
                    ws0.append([datetime.datetime(9999, 12, 31),name,"TRUE",name,"00:00:00",datetime.datetime(1900, 1, 1),unit,str(lt)+":00:00.00"])
                    for k in self.Q_Commodities_D_CO[c]:
                        if self.Q_Commodities_Segment[k][1] == p:
                            ws1.append(["FALSE",name,k])
                    r += 1
                    ws2.append([c + " [WH]",datetime.datetime(9999, 12, 31),"FALSE","FALSE","TRUE",name,r,dp,"0:00:00.00",datetime.datetime(1900, 1, 1),str(lt)+":00:00.00"])

            # NDP - CO lanes (reg procurement)
            for ndp,c in self.Q_Transport_NDP2CO: # lane = (NDP,CO)
                if (ndp,c) in self.Transport_Dur.keys():
                    lt = self.Transport_Dur[ndp,c]*24
                else:
                    lt = 40*24 # if the lead time is 0 I can't use the lane :(
                    self.weird.append(("No lead time found for lane",ndp,c))
                projects = [p for (co,p) in self.Q_SalesSegments if co == c] # Identify relevant projects
                for p in projects:
                    name = "Truck from " + ndp + " to " + c + " for " + p
                    ws0.append([datetime.datetime(9999, 12, 31),name,"TRUE",name,"00:00:00",datetime.datetime(1900, 1, 1),"Road - Overland",str(lt)+":00:00.00"])
                    for k in self.Q_Commodities_D_CO[c]:
                        sk = self.Q_Commodities_SpecCom[k]
                        if sk in self.Q_SpecificCommodities_S_NDP[ndp]:
                            if self.Q_Commodities_Segment[k][1] == p:
                                ws1.append(["FALSE",name,k])
                    r += 1
                    ws2.append([c + " [WH]",datetime.datetime(9999, 12, 31),"FALSE","FALSE","TRUE",name,r,ndp,"0:00:00.00",datetime.datetime(1900, 1, 1),str(lt)+":00:00.00"])

            # NDP - CO lanes (loc procurement)
            for ndp in self.Q_NDPs:
                c = self.MD_Locations_Country[ndp]
                if c in self.Q_COs:
                    if ndp in self.Q_DPs:
                        continue # link already established through DP2CO
                    projects = [p for (co,p) in self.Q_SalesSegments if co == c] # Identify relevant projects
                    for p in projects:
                        name = "Truck from " + ndp + " to " + c + " for " + p
                        ws0.append([datetime.datetime(9999, 12, 31),name,"TRUE",name,"00:00:00",datetime.datetime(1900, 1, 1),"Road - Inland","24:00:00.00"])
                        for k in self.Q_Commodities_D_CO[c]:
                            sk = self.Q_Commodities_SpecCom[k]
                            if sk in self.Q_SpecificCommodities_S_NDP[ndp]:
                                if self.Q_Commodities_Segment[k][1] == p:
                                    ws1.append(["FALSE",name,k])
                        r += 1
                        ws2.append([c + " [WH]",datetime.datetime(9999, 12, 31),"FALSE","FALSE","TRUE",name,r,ndp,"0:00:00.00",datetime.datetime(1900, 1, 1),"0:00:00.00"])

            # Distribution lanes
            for c in self.Q_COs:
                projects = [p for (co,p) in self.Q_SalesSegments if co == c] # Identify relevant projects
                for p in projects:
                    name = "Distribute in " + c + " for " + p
                    ws0.append([datetime.datetime(9999, 12, 31),name,"TRUE",name,"00:00:00",datetime.datetime(1900, 1, 1),"DISTRIBUTION","768:00:00.00"])
                    for k in self.Q_Commodities_D_CO[c]:
                        if self.Q_Commodities_Segment[k][1] == p:
                            ws1.append(["FALSE",name,k])
                    ws2.append([c + " [D]",datetime.datetime(9999, 12, 31),"FALSE","FALSE","TRUE",name,r,c + " [WH]","0:00:00.00",datetime.datetime(1900, 1, 1),"0:00:00.00"])

            wb.save(filename = path1)
            t_e = time.time()
            print "Duration: ", "{0:.3f}".format(t_e-t_s), " seconds"

        def SalesSegments():
            print "> MP_SalesSegments"
            t_s = time.time()
            path0 = os.path.join(self.data_dir,"MP_SalesSegments.xlsx")
            path1 = os.path.join(self.dest_dir,"MP_SalesSegments.xlsx")
            wb = xl.load_workbook(filename = path0, read_only=False)

            ws = wb.get_sheet_by_name("SalesSegmentSpecifications_MP")
            # row = ExpectedBudget	Name	PeriodAlignment	SupportCostPercentage  ConfirmedBudget
            budget_f = {}
            for (c,p,d,sk),(f,nf,t) in self.Q_Demand.items():
                if (c,p) not in budget_f.keys():
                    budget_f[c,p] = 0
                budget_f[c,p] += f * self.StaticFCR[c,p,sk] # summing over date and commodity

            T = 12 # length of time horizon
            for t in range(T):
                d = add_months(self.Q_Start_Planning,t)
                for c,p in self.Q_SalesSegments:
                    rb = self.MD_Projects_RB[p]
                    name = rb + " | " + c + " | " + p
                    if (c,p) in self.DSC.keys():
                        dsc = self.DSC[c,p]*100
                    else:
                        dsc = 20
                        self.weird.append(("DSC missing",c,p))
                    ws.append([budget_f[c,p]/T*2,name,d,dsc,""])
            # NB: this is just to test - now every project can spend at most X its average monthly budget in one month

            discounting = {}
            for c,p,sk in self.Q_Supply.keys():
                if (c,p) not in discounting.keys():
                    if (c,p) in self.Q_BudgetBonus.keys():
                        discounting[c,p] = -1 * self.Q_BudgetBonus[c,p]
                    else:
                        discounting[c,p] = 0
                discounting[c,p] += self.Q_Supply[c,p,sk] * self.StaticFCR[c,p,sk]

            ws = wb.get_sheet_by_name("SalesSegments_MP")
            # row = DisplayIndex	HasNetDemand	Name	ParentName	PriorityName	ExpectedBudget
            r = -1
            if self.Q_RBs == []: # one RB
                r += 1
                ws.append([r,"TRUE",self.Q_RB,"","None",""])
            else:
                for rb in self.Q_RBs:
                    r += 1
                    ws.append([r,"TRUE",rb,"","None",""])
            for c in self.Q_COs:
                r += 1
                rb = self.MD_Countries_RB[c]
                ws.append([r,"TRUE",rb + " | " + c,rb,"None",""])
            for c,p in self.Q_SalesSegments:
                r += 1
                rb = self.MD_Countries_RB[c]
                l1 = rb + " | " + c
                l2 = l1 + " | " + p
                if (c,p) in self.Project_Priority.keys():
                    priority = self.Project_Priority[c,p]
                else:
                    priority = "L1"
                    self.weird.append(("No Project Priority detected for",c,p))
                if (c,p) in self.DSC.keys():
                    dsc = self.DSC[c,p]
                else:
                    dsc = .2
                    self.weird.append(("DSC missing",c,p))
                if (c,p) in discounting.keys():
                    discount = discounting[c,p]
                else:
                    discount = 0
                exp_FCR = budget_f[c,p] - discount # discount available budget with stocks
                if exp_FCR < 0:
                    self.weird.append(("More discounting cost than available budget",c,p))
                    exp_FCR = 0
                exp_DOC = exp_FCR / 1.07 / (1 + dsc) # discount available budget with DSC and ISC
                ws.append([r,"TRUE",l2,l1,priority,exp_DOC])
                r += 1
                ws.append([r,"TRUE",l2+" [F]",l2,priority,""])
                r += 1
                ws.append([r,"FALSE",l2+" [NF]",l2,priority,""])

            ws = wb.get_sheet_by_name("SalesSegmentBudgets_MP")
            # row = ConfirmedBudget	End	Name	Start
            for c,p in self.Q_Budget.keys():
                usd = self.Q_Budget[c,p]
                if usd < 0:
                    usd = 0
                rb = self.MD_Projects_RB[p]
                ss = rb + " | " + c + " | " + p
                ws.append([usd,add_months(self.Q_Start_Planning,1),ss,self.Q_Start_Planning])

            wb.save(filename = path1)
            t_e = time.time()
            print "Duration: ", "{0:.3f}".format(t_e-t_s), " seconds"

        def SalesDemands():
            print "> MP_SalesDemands"
            t_s = time.time()
            path0 = os.path.join(self.data_dir,"MP_SalesDemands.xlsx")
            path1 = os.path.join(self.dest_dir,"MP_SalesDemands.xlsx")
            wb = xl.load_workbook(filename = path0, read_only=False)
            ws = wb.get_sheet_by_name("SalesDemands_MP")
            # row = CanBePostponed	CurrencyID	EndDate	ID	PriorityName	ProductID
            # Quantity	Revenue	SalesSegmentName	StartDate	StockingPointID	UnitOfMeasureName	DemandUncertaintyPercentage

            # Actual demand
            r = 0
            for (c,p,d,sk),(f,nf,t) in self.Q_Demand.items():
                kp = sk + "_" + c + "_" + p
                gk = kp + " [GMO]"
                if gk not in self.Q_Commodities:
                    k = kp + " [Non GMO]"
                elif self.GMO[c] == 0:
                    k = kp + " [Non GMO]"
                else:
                    k = kp
                if f > 0:
                    rb = self.MD_Projects_RB[p]
                    name = rb + " | " + c + " | " + p + " [F]"
                    r += 1
                    ws.append(["FALSE","USD",add_months(d,1),r,"High",k,max(f,0.1),0,name,d,c + " [D]","MT",0])
                if nf > 0:
                    rb = self.MD_Projects_RB[p]
                    name = rb + " | " + c + " | " + p + " [NF]"
                    r += 1
                    ws.append(["FALSE","USD",add_months(d,1),r,"None",k,max(nf,0.1),0,name,d,c + " [D]","MT",0])
            wb.save(filename = path1)

            # Dummy demand
            path1 = os.path.join(self.dest_dir,"Dummy\MP_SalesDemands.xlsx")
            wb = xl.load_workbook(filename = path0, read_only=False)
            ws = wb.get_sheet_by_name("SalesDemands_MP")
            # row = CanBePostponed	CurrencyID	EndDate	ID	PriorityName	ProductID
            # Quantity	Revenue	SalesSegmentName	StartDate	StockingPointID	UnitOfMeasureName	DemandUncertaintyPercentage
            r = 0
            for c,p in self.Q_SalesSegments:
                rb = self.MD_Projects_RB[p]
                name = rb + " | " + c + " | " + p + " [F]"
                for k in self.Q_Commodities_D_CO[c]:
                    if self.Q_Commodities_Segment[k] != (c,p):
                        continue
                    for t in range(12):
                        d = add_months(self.Q_Start_Planning,t)
                        r += 1
                        ws.append(["FALSE","USD",add_months(d,1),r,"High",k,100,0,name,d,c + " [D]","MT",0])

            wb.save(filename = path1)
            t_e = time.time()
            print "Duration: ", "{0:.3f}".format(t_e-t_s), " seconds"

        def EntityCosts():
            print "> MP_EntityCosts"
            t_s = time.time()
            path0 = os.path.join(self.data_dir,"MP_EntityCosts.xlsx")
            path1 = os.path.join(self.dest_dir,"MP_EntityCosts.xlsx")
            wb = xl.load_workbook(filename = path0, read_only=False)

            ws = wb.get_sheet_by_name("LaneCosts_MP")
            # row = AccountName	Cost	CostDriver	ID	LaneID	LengthOfTime	Start	TimeUnit
            r = 0
            # NDP - DP lanes (Shipping)
            conv = {}
            for t in self.MD_ShippingTypes:
                conv[t] = (1-self.Bulk_Discount[t]) / 22 # conversion: x% discount, 22 MT/container avg
            conv["Container"] = 1
            for lane in self.Q_Shipping:
                for t in self.MD_ShippingTypes:
                    lp,dp = lane[0],lane[1]
                    name = "Ship from " + lp + " to " + dp + " in " + t
                    if lane in self.Shipping_Rate.keys():
                        sr = self.Shipping_Rate[lane] * conv[t]
                    else:
                        sr = 0
                        self.weird.append(("Missing shipping rate",lane))
                    pc = 100 if t != "Container" else 2200
                    # appropriate cost tbd!! 100 is pretty good performance, but still some direct shipments occurring - investigate
                    if t != "Bulk":
                        r += 1
                        ws.append(["Shipping cost",sr,"Volume",r,name,1,self.Q_Start_Horizon,"Day"])
                        if dp not in self.GCMF_Ports and lp not in self.GCMF_Ports and self.GCMF_Priority=="Soft":
                            r += 1
                            ws.append(["Process preference cost",pc,"Volume",r,name,1,self.Q_Start_Horizon,"Day"])
                    else:
                        if (lp,dp) not in self.Q_Bulk_Coms.keys():
                            continue
                        for (sk,gmo) in self.Q_Bulk_Coms[lp,dp]:
                            gmo_s = " [GMO]" if gmo == 1 else " [Non GMO]"
                            name_b = name + " (" + sk + gmo_s + ")"
                            r += 1
                            ws.append(["Shipping cost",sr,"Volume",r,name_b,1,self.Q_Start_Horizon,"Day"])
                            if dp not in self.GCMF_Ports and lp not in self.GCMF_Ports and self.GCMF_Priority=="Soft":
                                r += 1
                                ws.append(["Process preference cost",pc,"Volume",r,name_b,1,self.Q_Start_Horizon,"Day"])

            # DP - CO lanes (overland/inland transport)
            for dp,c in self.Q_Transport_DP2CO:
                projects = [p for (co,p) in self.Q_SalesSegments if co == c] # Identify relevant projects
                if self.MD_Locations_Country[dp] != c: # Overland
                    if (dp,c) in self.Transport_Cost.keys():
                        otc = self.Transport_Cost[dp,c]
                    else:
                        otc = 50
                        self.weird.append(("Missing Overland Transport cost",dp,c))
                else: # Inland
                    otc = ""
                for p in projects:
                    name = "Truck from " + dp + " to " + c + " for " + p
                    if (c,p) in self.LTSH.keys():
                        ltsh = self.LTSH[c,p]["Overseas"] # Overseas rate
                        if ltsh == (0,0):
                            self.weird.append(("Missing Overseas LTSH rate",c,p))
                            ltsh = self.LTSH[c,p]["Average"] # Average rate
                    else:
                        ltsh = (0,0)
                        self.weird.append(("Missing LTSH rates",c,p))
                    if otc == "": # Inland
                        r += 1
                        ws.append(["LTSH",sum(ltsh),"Volume",r,name,1,self.Q_Start_Horizon,"Day"]) # OVL + ITSH
                    else: # Overland
                        r += 1
                        ws.append(["LTSH",ltsh[1],"Volume",r,name,1,self.Q_Start_Horizon,"Day"]) # only ITSH
                        r += 1
                        ws.append(["Overland transport cost",otc,"Volume",r,name,1,self.Q_Start_Horizon,"Day"])

            # NDP - CO lanes (reg procurement)
            for ndp,c in self.Q_Transport_NDP2CO:
                if (ndp,c) in self.Transport_Cost.keys():
                    otc = self.Transport_Cost[ndp,c]
                else:
                    otc = 50
                    self.weird.append(("Missing Overland Transport cost",dp,c))
                projects = [p for (co,p) in self.Q_SalesSegments if co == c] # Identify relevant projects
                for p in projects:
                    name = "Truck from " + ndp + " to " + c + " for " + p
                    if (c,p) in self.LTSH.keys():
                        ltsh = self.LTSH[c,p]["Regional"] # Regional rate
                        if ltsh == (0,0):
                            self.weird.append(("Missing Regional LTSH rate",c,p))
                            ltsh = self.LTSH[c,p]["Average"] # Average rate
                    else:
                        ltsh = (0,0)
                        self.weird.append(("Missing LTSH rates",c,p))
                    r += 1
                    ws.append(["LTSH",ltsh[1],"Volume",r,name,1,self.Q_Start_Horizon,"Day"]) # only ITSH
                    r += 1
                    ws.append(["Overland transport cost",otc,"Volume",r,name,1,self.Q_Start_Horizon,"Day"])

            # NDP - CO lanes (loc procurement)
            for ndp in self.Q_NDPs:
                c = self.MD_Locations_Country[ndp]
                if c in self.Q_COs:
                    if ndp in self.Q_DPs:
                        continue # link already established through DP2CO
                    projects = [p for (co,p) in self.Q_SalesSegments if co == c] # Identify relevant projects
                    for p in projects:
                        name = "Truck from " + ndp + " to " + c + " for " + p
                        if (c,p) in self.LTSH.keys():
                            ltsh = self.LTSH[c,p]["Local"] # Local rate
                            if ltsh == (0,0):
                                self.weird.append(("Missing Local LTSH rate",c,p))
                                ltsh = self.LTSH[c,p]["Average"] # Average rate
                        else:
                            ltsh = (0,0)
                            self.weird.append(("Missing LTSH rates",c,p))
                        r += 1
                        ws.append(["LTSH",sum(ltsh),"Volume",r,name,1,self.Q_Start_Horizon,"Day"]) # ITSH + OVL

            # Distribution lanes
            for c in self.Q_COs:
                projects = [p for (co,p) in self.Q_SalesSegments if co == c] # Identify relevant projects
                for p in projects:
                    name = "Distribute in " + c + " for " + p
                    if (c,p) in self.ODOC.keys():
                        odoc = self.ODOC[c,p]
                    else:
                        odoc = 50
                        self.weird.append(("Missing ODOC rate",c,p))
                    r += 1
                    ws.append(["ODOC",odoc,"Volume",r,name,1,self.Q_Start_Horizon,"Day"])

            ws = wb.get_sheet_by_name("OperationCosts_MP")
            # row = AccountName	Cost	CostDriver	ID	LengthOfTime	OperationID	Start	TimeUnit
            r = 0
            # Procurement costs
            for key in self.Q_Procurement: # key = (oc,ndp,com,gmo,date)
                oc,ndp,sk,gmo,d0 = key[0],key[1],key[2],key[3],key[4]
                if gmo == 1:
                    gmo_s = " [GMO]"
                else:
                    gmo_s = " [Non GMO]"
                rname = "Buy " + sk + gmo_s + " from " + oc + " at " + ndp
                if d0.day < 15:
                    d = datetime.datetime(d0.year,d0.month,1)
                else:
                    d1 = add_months(d0,1)
                    d = datetime.datetime(d1.year,d1.month,1)
                # NB: in MP if a price is set for 5 august, the new price is active from 1 sep onwards -> this is a bit more accurate
                coms = [k for k in self.Q_Commodities_D_NDP[ndp] if self.Q_Commodities_SpecCom[k] == sk and self.Q_Commodities_GMO[k] == gmo]
                if ndp in self.Q_DPs:
                    for k in self.Q_Commodities_D_DP[ndp]:
                        if self.Q_Commodities_SpecCom[k] == sk and self.Q_Commodities_GMO[k] == gmo:
                            coms.append(k)
                coms = list(set(coms))
                for k in coms:
                    oname = "Buy " + k + " from " + oc + " at " + ndp
                    r += 1
                    ws.append(["Sourcing cost",self.Procurement_Cost[key],"Volume",r,1,oname,d,"Day"])
                    if d > self.Q_Start_Horizon:
                        # NB: some options don't have a recent procurement price, but get one during the time horizon
                        # this fills the month gap until that actual price with a historic (or phantom) price
                        aod = self.Procurement_Date[key]
                        r += 1
                        if aod <= self.Q_Start_Horizon: # use as of date to fill data gap
                            ws.append(["Sourcing cost",self.Procurement_Cost[key],"Volume",r,1,oname,aod,"Day"])
                        else: # find historic option or add phantom data
                            all_options = [proc for proc in self.Procurement_Cost.keys() if proc[0] == oc and proc[1] == ndp and proc[2] == sk and proc[3] == gmo]
                            dmax = datetime.datetime(1900,1,1)
                            for option in all_options: # option = (oc,ndp,com,gmo,date)
                                d_u = option[4] # last updated date
                                d_a = self.Procurement_Date[option] # as of date
                                if d_a <= self.Q_Start_Horizon and d_u > dmax:
                                    dmax = d_u
                            if dmax > datetime.datetime(1900,1,1): # we found a good historic data point
                                ws.append(["Sourcing cost",self.Procurement_Cost[oc,ndp,sk,gmo,dmax],"Volume",r,1,oname,self.Procurement_Date[oc,ndp,sk,gmo,dmax],"Day"])
                            else: # we didn't find a good historic point -> phantom data
                                ws.append(["Sourcing cost",9999,"Volume",r,1,oname,self.Q_Start_Horizon,"Day"])
                            # NB: we could do this for every month between Q_Start_Horizon and Q_Start_Planning to get a credible data set going
                            # the merits are limited though, seeing as only the most recent price will be used anyway
                            # also take heed that once we add forecasts this algorithm may have to be adapted!
            # Price Forecasts
            for key in self.Q_Forecast: # key = (oc,ndp,com,gmo,date)
                oc,ndp,sk,gmo,d = key[0],key[1],key[2],key[3],key[4]
                if gmo == 1:
                    gmo_s = " [GMO]"
                else:
                    gmo_s = " [Non GMO]"
                rname = "Buy " + sk + gmo_s + " from " + oc + " at " + ndp
                coms = [k for k in self.Q_Commodities_D_NDP[ndp] if self.Q_Commodities_SpecCom[k] == sk and self.Q_Commodities_GMO[k] == gmo]
                if ndp in self.Q_DPs:
                    for k in self.Q_Commodities_D_DP[ndp]:
                        if self.Q_Commodities_SpecCom[k] == sk and self.Q_Commodities_GMO[k] == gmo:
                            coms.append(k)
                coms = list(set(coms))
                for k in coms:
                    oname = "Buy " + k + " from " + oc + " at " + ndp
                    r += 1
                    ws.append(["Sourcing cost",self.Forecast[key],"Volume",r,1,oname,d,"Day"])
            # In-Kind 'procurement'
            check = []
            for don,sk,o,aod,dur,cor,c,p in self.Q_Donation:
                rname = "Donation of " + sk + " by " + don + " at " + o
                k = sk + "_" + c + "_" + p + " [Non GMO]"
                oname = "Donation of " + k + " by " + don + " at " + o
                if oname in check:
                    continue
                if (don,o,sk) in self.IK_Price.keys():
                    cost = self.IK_Price[don,o,sk][0]
                    r += 1
                    ws.append(["Sourcing cost",cost,"Volume",r,1,oname,self.Q_Start_Horizon,"Day"])
                    check.append(oname)
                else:
                    self.weird.append(("Missing cost for donation",don,o,sk))
            # GCMF procurement
            for loc,sk in self.Q_GCMF:
                rname = "Buy " + sk + " from GCMF at " + loc
                coms = [k for k in self.Q_Commodities if self.Q_Commodities_SpecCom[k] == sk and self.Q_Commodities_GMO[k] == 0]
                for k in coms:
                    oname = "Buy " + k + " from GCMF at " + loc
                    c,p = self.Q_Commodities_Segment[k]
                    if loc in self.Q_COs and loc != c:
                        continue
                    cost = self.Q_GCMF_Price[sk][0]
                    r += 1
                    ws.append(["Sourcing cost",cost,"Volume",r,1,oname,self.Q_Start_Horizon,"Day"])
                    if loc in self.Q_COs: # stocks already in the CO
                        ltsh = self.LTSH[loc,p]["Local"]
                        r += 1
                        ws.append(["LTSH",sum(ltsh),"Volume",r,1,oname,self.Q_Start_Horizon,"Day"])
##                    # preference cost
##                    for t in range(10):
##                        d = add_months(self.Q_Start_Planning,t)
##                        bonus = 10 - t
##                        r += 1
##                        ws.append(["Process preference bonus",bonus,"Volume",r,1,oname,d,"Day"]) # preference cost
            # Unloading costs
            for dp in self.Q_DPs:
                for k in self.Q_Commodities_D_DP[dp]:
                    sk,gmo = self.Q_Commodities_SpecCom[k], self.Q_Commodities_GMO[k]
                    if gmo == 1:
                        gmo_s = " [GMO]"
                    else:
                        gmo_s = " [Non GMO]"
                    for t in self.MD_ShippingTypes:
                        dpt = dp + " [" + t + "]"
                        rname = "Unloading " + sk + gmo_s + " in " + dpt
                        oname = "Unloading " + k + " in " + dpt
                        if (dp,t) in self.Handling_Cost.keys():
                            hc = self.Handling_Cost[dp,t]
                        else:
                            hc = 20
                            self.weird.append(("Missing handling cost",dp,t))
                        r += 1
                        ws.append(["Handling cost",hc,"Volume",r,1,oname,self.Q_Start_Horizon,"Day"])
##                        if dp in self.GCMF_Ports:
##                            for t in range(10):
##                                d = add_months(self.Q_Start_Planning,t)
##                                r += 1
##                                ws.append(["Process preference bonus",50,"Volume",r,1,oname,d,"Day"])
##                                # appropriate bonus value tbd

            # Storage costs
            ws = wb.get_sheet_by_name("ProductValueAndCosts_MP")
            # row = AccountName	Cost	CostDriver	ID	ProductID	Start
            r = 0
            for k in self.Q_Commodities:
                r += 1
                ws.append(["Storage cost",60,"Inventory holding",r,k,self.Q_Start_Horizon])
                # NB: flat rate of 5 $/mt/month (=60 $/mt/year) for each commodity as baseline
            for sk in self.Q_GCMF_Commodities:
                k = sk + "_GCMF [Non GMO]"
                r += 1
                ws.append(["Storage cost",50,"Inventory holding",r,k,self.Q_Start_Horizon])
            # PISP costs tab can be used to override the flat rate

            wb.save(filename = path1)
            t_e = time.time()
            print "Duration: ", "{0:.3f}".format(t_e-t_s), " seconds"

        def InventorySupplies():
            print "> MP_InventorySupplies"
            t_s = time.time()
            path0 = os.path.join(self.data_dir,"MP_InventorySupplies.xlsx")

            for baseline in [0,1]:
                if baseline == 0:
                    path1 = os.path.join(self.dest_dir,"MP_InventorySupplies.xlsx")
                else:
                    path1 = os.path.join(self.dest_dir,"Baseline\MP_InventorySupplies.xlsx")
                wb = xl.load_workbook(filename = path0, read_only=False)
                ws = wb.get_sheet_by_name("InventorySupplies_MP")
                # row = Date	DeadInventoryQuantity	Description	ID	ProductID	StockingPointID	UserQuantity

                r = 0
                for key in self.Q_Inventory:
                    loc,sk,c,p = key[0],key[1],key[2],key[3]
                    kp = sk + "_" + c + "_" + p
                    k = kp + " [Non GMO]"
                    inv,itr = self.Inventory[key]
                    d = self.Q_Start_Planning
                    if loc in self.Q_COs: # split up inventory over [WH] and [D]
                        try:
                            q = self.Q_Demand[loc,p,d,sk][0]
                        except:
                            q = 0 # no demand in first month
                        if inv > 0: # initial inventory
                            if q > 0:
                                if inv <= q: # inventory is less than first month of demand
                                    r += 1
                                    ws.append([d,0,"",r,k,loc + " [D]",max(inv,0.1)])
                                else: # inventory is more than first month of demand
                                    r += 1
                                    ws.append([d,0,"",r,k,loc + " [D]",max(q,0.1)])
                                    r += 1
                                    ws.append([d,0,"",r,k,loc + " [WH]",max(inv-q,0.1)])
                            else: # no immediate demand
                                r += 1
                                ws.append([d,0,"",r,k,loc + " [WH]",max(inv,0.1)])
                        if itr > 0: # in transit stocks
                            r += 1
                            ws.append([d,0,"",r,k,loc + " [WH]",max(itr,0.1)])
                    else: # loc is a DP
                        if inv > 0:
                            r += 1
                            ws.append([d,0,"",r,k,loc,max(inv,0.1)])
                        if itr > 0:
                            r += 1
                            ws.append([add_months(d,1),0,"",r,k,loc,max(itr,0.1)])

                for loc,sk in self.Q_GCMF:
                    k = sk + "_GCMF [Non GMO]"
                    d1 = self.Q_Start_Planning
                    d2 = add_months(d1,1)
                    if (self.Q_RB,loc,sk) not in self.GCMF.keys():
                        continue # currently no inventory
                    inv,itr = self.GCMF[self.Q_RB,loc,sk]
                    if loc in self.Q_COs:
                        if inv > 0:
                            r += 1
                            ws.append([d1,0,"",r,k,loc + " [WH]",max(inv,0.1)])
                        if itr > 0:
                            r += 1
                            ws.append([d2,0,"",r,k,loc + " [WH]",max(itr,0.1)])
                    else: # loc in Q_DPs
                        if inv > 0:
                            r += 1
                            ws.append([d1,0,"",r,k,loc,max(inv,0.1)])
                        if itr > 0:
                            r += 1
                            ws.append([d2,0,"",r,k,loc,max(itr,0.1)])

                if baseline == 0:
                    for key in self.Q_Donation:
                        r += 1
                        don,sk,aod,c,p = key[0],key[1],key[3],key[6],key[7]
                        qty = self.Donation[key]
                        if qty < 0.1:
                            continue
                        if aod < self.Q_Start_Planning:
                            aod = self.Q_Start_Planning
                        k = sk + "_" + c + "_" + p + " [Non GMO]"
                        ws.append([aod,0,don,r,k,"In-Kind Donation",qty])
                else:
                    reqs = {}
                    for (c,p,d,sk),(f,nf,t) in self.Q_Demand.items():
                        if (c,sk) not in reqs.keys():
                            reqs[c,sk] = {}
                        try:
                            reqs[c,sk][p] += f
                        except:
                            reqs[c,sk][p] = f
                    check = []
                    for (rc,sk,ic) in self.allocation.keys():
                        if (rc,sk) not in reqs.keys():
                            continue
                        if (rc,sk) not in check:
                            check.append((rc,sk))
                        wh = "Virtual Warehouse [" + ic + "]"
                        for p in reqs[rc,sk].keys():
                            k = sk + "_" + rc + "_" + p + " [Non GMO]"
                            req = reqs[rc,sk][p]
                            sup = self.Q_Supply[rc,p,sk] if (rc,p,sk) in self.Q_Supply.keys() else 0
                            mt = req - sup
                            if mt <= 0: # more supply than demand
                                continue
                            q = mt * self.allocation[rc,sk,ic]
                            r += 1
                            ws.append([self.Q_Start_Planning,0,"",r,k,wh,max(q,0.1)])
                    for (rc,sk) in (set(reqs.keys())-set(check)): # commodities with no purchasing history
                        ics = [key[2] for key in self.allocation.keys() if key[1]==sk] # ICs in use for this commodity by other COs
                        ics = list(set(ics))
                        for ic in ics:
                            if ic not in self.Q_Countries: # not relevant for
                                continue
                            wh = "Virtual Warehouse [" + ic + "]"
                            for p in reqs[rc,sk].keys():
                                k = sk + "_" + rc + "_" + p + " [Non GMO]"
                                q = reqs[rc,sk][p]
                                r += 1
                                ws.append([self.Q_Start_Planning,0,"",r,k,wh,max(q,0.1)])

                for rc,p,po,sk,gmo,oc,ndp,d in self.Q_OpenPOs:
                    pod = add_months(d,-2)
                    if diff_month(self.Q_Start_Planning,pod) <= 0: # PO starts during planning horizon
                        continue
                    if p.startswith("S"):
                        k = sk + "_GCMF [Non GMO]"
                    else:
                        gmo_s = " [GMO]" if gmo==1 else " [Non GMO]"
                        k = sk + "_" + rc + "_" + p + gmo_s
                    aod = max(d,self.Q_Start_Planning)
                    #NB: POs that should have arrived in the last 2 months are estimated to arrive during the first month of the planning horizon
                    r += 1
                    ws.append([aod,0,po,r,k,ndp,max(self.OpenPOs[rc,p,po,sk,gmo,oc,ndp,d],0.1)])

                # Virtual Warehouse (BBD)
                for c,p,sk,d in self.Q_BBDinv.keys():
                    q = self.Q_BBDinv[c,p,sk,d]
                    k = sk + "_" + c + "_" + p + " [Non GMO]"
                    r += 1
                    ws.append([d,0,"",r,k,"Virtual Warehouse [BBD]",max(q,0.1)])

                wb.save(filename = path1)

            # Dummy inventory
            path1 = os.path.join(self.dest_dir,"Dummy\MP_InventorySupplies.xlsx")
            wb = xl.load_workbook(filename = path0, read_only=False)
            ws = wb.get_sheet_by_name("InventorySupplies_MP")
            # row = Date	DeadInventoryQuantity	Description	ID	ProductID	StockingPointID	UserQuantity
            r = 0
            for oname in self.Q_Operations:
                k,vk,ndp,t = self.Q_Operations[oname]
                r += 1
                ws.append([self.Q_Start_Planning,0,oname,r,k,"Virtual Warehouse [BBD]",12])
                if t == "Procurement":
                    ic = self.MD_Locations_Country[ndp]
                    wh = "Virtual Warehouse [" + ic + "]"
                    r += 1
                    ws.append([self.Q_Start_Planning,0,oname,r,vk,wh,12])
                elif t == "GCMF":
                    r += 1
                    ws.append([self.Q_Start_Planning,0,oname,r,vk,ndp,12])
                else: # t == "Donation"
                    ic = self.MD_Locations_Country[ndp]
                    wh = "Virtual Warehouse [" + ic + "]"
                    r += 1
                    ws.append([self.Q_Start_Planning,0,oname,r,vk,wh,12])
                    r += 1
                    ws.append([self.Q_Start_Planning,0,oname,r,vk,"In-Kind Donation",12])

            wb.save(filename = path1)
            t_e = time.time()
            print "Duration: ", "{0:.3f}".format(t_e-t_s), " seconds"

        def StockingPointCapacities():
            print "> MP_StockingPointCapacities"
            t_s = time.time()
            path0 = os.path.join(self.data_dir,"MP_StockingPointCapacities.xlsx")
            path1 = os.path.join(self.dest_dir,"MP_StockingPointCapacities.xlsx")
            wb = xl.load_workbook(filename = path0, read_only=False)
            ws = wb.get_sheet_by_name("StockingPointCapacities_MP")
            # row = MaxCapacity	Start	StockingPointID

            for dp in self.Q_DPs:
                for t in self.MD_ShippingTypes:
                    dpt = dp + " [" + t + "]"
                    ws.append([0,self.Q_Start_Horizon,dpt])
                if dp in self.Storage_Capacity.keys():
                    cap = self.Storage_Capacity[dp]
                else:
                    self.weird.append(("No capacity known",dp))
                    continue # no capacity known --> setting to 0 gets messy when adding existing inventories
                ws.append([cap,self.Q_Start_Horizon,dp])
            for ndp in self.Q_NDPs:
                if ndp in self.Q_DPs:
                    continue
                ws.append([0,self.Q_Start_Horizon,ndp])
            for co in self.Q_COs:
                ws.append([0,self.Q_Start_Horizon,co + " [D]"])
                if co in self.Storage_Capacity_Country.keys():
                    cap = self.Storage_Capacity_Country[co]
                else:
                    self.weird.append(("No capacity known",co))
                    continue # no capacity known --> setting to 0 gets messy when adding existing inventories
                ws.append([cap,self.Q_Start_Horizon,co + " [WH]"])

            wb.save(filename = path1)
            t_e = time.time()
            print "Duration: ", "{0:.3f}".format(t_e-t_s), " seconds"

        def UnitCapacities():
            print "> MP_UnitCapacities"
            t_s = time.time()
            path0 = os.path.join(self.data_dir,"MP_UnitCapacities.xlsx")
            path1 = os.path.join(self.dest_dir,"MP_UnitCapacities.xlsx")
            wb = xl.load_workbook(filename = path0, read_only=False)
            ws = wb.get_sheet_by_name("UnitCapacities_MP")
            # row = HasUserMaximumLoadPercentage	LengthOfTime	MaxCapacity	MaximumLoadPercentage	MinCapacity
            # NrOfUnitsOpen	NrOfUnitsTemporarilyClosed	Start	TimeUnit	UnitID

            for key in self.Handling_Capacity.keys(): # key = (loc,type,date)
                loc,t,d = key[0],key[1],key[2]
                if loc in self.Q_DPs:
                    dpt = loc + " [" + t + "]"
                    cap = self.Handling_Capacity[key]
                    if cap == None:
                        cap = 0
                    ws.append(["FALSE",1,cap,100,0,1,0,d,"Month",dpt])

            wb.save(filename = path1)
            t_e = time.time()
            print "Duration: ", "{0:.3f}".format(t_e-t_s), " seconds"

        def InventorySpecifications():
            print "> MP_InventorySpecifications"
            t_s = time.time()
            path0 = os.path.join(self.data_dir,"MP_InventorySpecifications.xlsx")
            for baseline in [0,1]:
                if baseline == 0:
                    path1 = os.path.join(self.dest_dir,"MP_InventorySpecifications.xlsx")
                else:
                    path1 = os.path.join(self.dest_dir,"Baseline\MP_InventorySpecifications.xlsx")

                wb = xl.load_workbook(filename = path0, read_only=False)
                ws = wb.get_sheet_by_name("InventorySpecifications_MP")
                # row = HasMaxLevel	HasMaxLevelInDays	HasMinLevelInDays	HasTargetInDays	MaxLevelInDays	MaxLevelInQuantity
                # MinLevelInDays	MinLevelInQuantity	ProductID	Start	StockingPointID	TargetInDays	TargetInQuantity

                # Safety Stocks (next month of demand)
##                ss = {}
##                for c,p,d,k in self.Q_Demand: # dem = (country, project, date, commodity)
##                    kp = k + "_" + c + "_" + p
##                    d2 = add_months(d,-2)
##                    if d2 < self.Q_Start_Planning:
##                        continue
##                    key = (c,kp,d2)
##                    if key in ss.keys():
##                        ss[key] += self.TacticalDemand_CF[c,p,d,k] + self.TacticalDemand_IK[c,p,d,k]
##                    else:
##                        ss[key] = self.TacticalDemand_CF[c,p,d,k] + self.TacticalDemand_IK[c,p,d,k]
##                for c in self.Q_COs:
##                    for k in self.Q_Commodities_D_CO[c]:
##                        for t in range(12):
##                            if self.Q_Commodities_GMO[k] == 1:
##                                continue
##                            kp = k[:-10]
##                            d = add_months(self.Q_Start_Planning,t)
##                            if (c,kp,d) in ss.keys():
##                                target = ss[c,kp,d]
##                            else:
##                                target = 0
##                            ws.append(["FALSE","FALSE","FALSE","FALSE",0,0,0,0,kp,d,c + " [WH]",0,target])

                # GCMF inventory
                for loc,sk in self.Q_GCMF:
                    k = sk + "_GCMF [Non GMO]"
                    d = add_months(self.Q_Start_Planning,6)
                    if loc in self.Q_COs:
                        ws.append(["TRUE","FALSE","FALSE","FALSE",0,0,0,0,k,d,loc + " [WH]",0,0])
                    else: # loc in Q_DPs
                        ws.append(["TRUE","FALSE","FALSE","FALSE",0,0,0,0,k,d,loc,0,0])
                for dp in self.Q_DPs:
                    if dp in self.GCMF_Ports:
                        for sk in self.Q_GCMF_Commodities:
                            if (dp,sk) not in self.Q_GCMF:
                                k = sk + "_GCMF [Non GMO]"
                                d = add_months(self.Q_Start_Planning,6)
                                ws.append(["TRUE","FALSE","FALSE","FALSE",0,0,0,0,k,d,dp,0,0])

                # Donations
                if baseline == 0:
                    ik_coms = []
                    for don,sk,o,aod,dur,cor,c,p in self.Q_Donation:
                        mt = self.Donation[don,sk,o,aod,dur,cor,c,p]
                        if mt < 0.1:
                            continue
                        if aod < self.Q_Start_Planning:
                            aod = self.Q_Start_Planning
                        k = sk + "_" + c + "_" + p + " [Non GMO]"
                        d = add_months(aod,dur-1)
                        ik_coms.append((k,d))
                    ik_coms = list(set(ik_coms))
                    for (k,d) in ik_coms:
                        ws.append(["TRUE","FALSE","FALSE","FALSE",0,0,0,0,k,d,"In-Kind Donation",0,0])
                wb.save(filename = path1)

            t_e = time.time()
            print "Duration: ", "{0:.3f}".format(t_e-t_s), " seconds"

        def Feedbacks():
            print "> MP_Feedbacks"
            t_s = time.time()
            path0 = os.path.join(self.data_dir,"MP_Feedbacks.xlsx")
            path1 = os.path.join(self.dest_dir,"MP_Feedbacks.xlsx")
            wb = xl.load_workbook(filename = path0, read_only=False)
            # NB: As a rule, feedbacks should always relate to actions
            # that start during the planning horizon, never before!
            # to model actions that started in the past, we have
            # to rely on InventorySupplies

            ws_o = wb.get_sheet_by_name("FeedbackPeriodTaskOperations_MP")
            # row = DateTime	Description	FeedbackQuantity	ID	OperationID
            ws_l = wb.get_sheet_by_name("FeedbackProductInTrips_MP")
            # row = Arrival	Description	DestinationStockingPointID	FeedbackQuantity	ID	LaneID	OriginStockingPointID	ProductID

            # Actual feedbacks
            r = 0
            for don,sk,o,aod,dur,cor,c,p in self.Q_Donation:
                k = sk + "_" + c + "_" + p + " [Non GMO]"
                qty = self.Donation[don,sk,o,aod,dur,cor,c,p]
                if qty < 0.1:
                    continue
                if aod < self.Q_Start_Planning:
                    aod = self.Q_Start_Planning
                descr = don if don != None else ""
                if dur > 1: # have some time to pick up the commodity
                    if cor == None: # corridor not specified
                        continue
                        # no feedbacks necessary
                    else: # corridor specified
                        self.weird.append(("Warning","Donation timing is flexible but corridor isn't - approach tbd"))
                        continue
                        # not sure if this will ever happen?
                else: # has to move immediately
                    if cor == None: # corridor not specified
                        continue
                        # no feedbacks necessary
                    else: # corridor specified
                        if (o,cor) not in self.Q_Shipping:
                            self.weird.append(("No shipping lane",o,cor))
                            continue
                        if (o,cor) in self.Shipping_Duration.keys():
                            m = int(round(self.Shipping_Duration[o,cor]/31))
                            arr = add_months(aod,m)
                            arr = datetime.datetime(arr.year,arr.month,1)
                        else:
                            self.weird.append(("No shipping lead time for donation",don,sk,o,cor))
                            continue
                        if cor not in self.GCMF_Ports and self.GCMF_Commodity[c,p,sk]==1 and self.GCMF_Priority=="Hard":
                            self.weird.append(("Infeasible corridor because of GCMF constraint",o,cor,sk))
                            continue
                        lane = "Ship from " + o + " to " + cor + " in Container"
                        # could consider alternative shipping types
                        r += 1
                        ws_l.append([arr,descr,cor + " [Container]",qty,r,lane,o,k])

            for rc,p,po,sk,gmo,oc,ndp,d in self.Q_OpenPOs:
                pod = add_months(d,-2)
                if diff_month(self.Q_Start_Planning,pod) > 0: # PO started before planning horizon
                    continue
                if p.startswith("S"):
                    self.weird.append(("Warning","Future GCMF PO - functionality not built yet"))
                    continue
                    # NB: This may happens when PO data is more recent than the planning horizon
                gmo_s = " [GMO]" if gmo==1 else " [Non GMO]"
                k = sk + "_" + rc + "_" + p + gmo_s
                o = "Buy " + k + " from " + oc + " at " + ndp
                r += 1
                ws_o.append([pod,po,self.OpenPOs[rc,p,po,sk,gmo,oc,ndp,d],r,o])
            wb.save(filename = path1)

            # Dummy feedbacks
            path1 = os.path.join(self.dest_dir,"Dummy\MP_Feedbacks.xlsx")
            wb = xl.load_workbook(filename = path0, read_only=False)
            ws = wb.get_sheet_by_name("FeedbackPeriodTaskOperations_MP")
            r = 0
            for oname in self.Q_Operations:
                for t in range(12):
                    d = add_months(self.Q_Start_Planning,t)
                    r += 1
                    ws.append([d,"",1,r,oname])

            wb.save(filename = path1)
            t_e = time.time()
            print "Duration: ", "{0:.3f}".format(t_e-t_s), " seconds"

        def SupplySpecifications():
            print "> MP_SupplySpecifications"
            t_s = time.time()
            path0 = os.path.join(self.data_dir,"MP_SupplySpecifications.xlsx")
            path1 = os.path.join(self.dest_dir,"MP_SupplySpecifications.xlsx")
            wb = xl.load_workbook(filename = path0, read_only=False)
            ws = wb.get_sheet_by_name("SupplySpecifications_MP")
            # row = End	HasMaxQuantity	MaxQuantity	MinQuantity	ProductID	Start	TargetQuantity	UnitID

            # Actual restrictions
            for oc,k0,c,p,d0,d1,ag in self.Q_Sourcing_Restrictions:
                cap = self.Sourcing_Restrictions[oc,k0,c,p,d0,d1,ag]
                if d1 == None:
                    d1 = add_months(self.Q_Start_Planning,12)
                if d0 < self.Q_Start_Horizon:
                    d0 = self.Q_Start_Horizon
                if p != None and c != None:
                    if k0 in self.Q_SpecificCommodities:
                        k = k0 + "_" + c + "_" + p
                    else:
                        self.weird.append(("Warning","Project-specific sourcing restrictions can only be defined at specific commodity level, not at commodity type level"))
                        continue
                else:
                    k = k0
                if ag == 1:
                    ws.append([d1,"TRUE",cap,0,k,d0,0,oc])
                else:
                    for t in range(diff_month(d1,d0)):
                        d_e = add_months(d0,t+1)
                        d_s = add_months(d0,t)
                        ws.append([d_e,"TRUE",cap,0,k,d_s,0,oc])

            wb.save(filename = path1)
            t_e = time.time()
            print "Duration: ", "{0:.3f}".format(t_e-t_s), " seconds"

        def Dummy():
            print "> Creating Dummy Data"
            t_s = time.time()
            same = ["EntityCosts","Lanes","Periods","Products","Routings","SalesSegments",\
                    "StockingPoints","UnitOfMeasures","Units"]
            dummy = os.path.join(self.dest_dir,"Dummy")
            for mp in same:
                mp = "MP_" + mp + ".xlsx"
                path0 = os.path.join(self.dest_dir,mp)
                path1 = os.path.join(dummy,mp)
                copyfile(path0,path1)
            t_e = time.time()
            print "Duration: ", "{0:.3f}".format(t_e-t_s), " seconds"

        def Baseline():
            print "> Creating Baseline Data"
            t_s = time.time()
            same = ["Periods","UnitOfMeasures","Units","StockingPoints","Products","Lanes","SalesSegments",\
                    "SalesDemands","EntityCosts","StockingPointCapacities","UnitCapacities"]
            bl = os.path.join(self.dest_dir,"Baseline")
            for mp in same:
                mp = "MP_" + mp + ".xlsx"
                path0 = os.path.join(self.dest_dir,mp)
                path1 = os.path.join(bl,mp)
                copyfile(path0,path1)
            t_e = time.time()
            print "Duration: ", "{0:.3f}".format(t_e-t_s), " seconds"

        # Execute
        self.weird = []
        Periods()
        UnitOfMeasures()
        Units()
        StockingPoints()
        Products()
        Routings()
        Lanes()
        SalesSegments()
        SalesDemands()
        EntityCosts()
        InventorySupplies()
        StockingPointCapacities()
        UnitCapacities()
        InventorySpecifications()
        Feedbacks()
        SupplySpecifications()
        Dummy()
        Baseline()
        print "> Exporting weird occurences to 'Debugging.xlsx'"
        self.print_to_file(self.weird,"Weird occurences (export)",["Type","Details"])

    def export_Optimus(self):
        None

    def draw_GUI(self,window):
        None

def add_months(sourcedate,months):
    month = sourcedate.month - 1 + months
    year = int(sourcedate.year + month / 12 )
    month = month % 12 + 1
    day = min(sourcedate.day,calendar.monthrange(year,month)[1])
    return datetime.datetime(year,month,day)

def diff_month(dbig, dsmall):
    return (dbig.year - dsmall.year)*12 + dbig.month - dsmall.month

MP_Mapping()
